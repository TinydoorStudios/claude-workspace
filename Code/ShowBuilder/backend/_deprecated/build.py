"""
build.py — MAC ONLY. Turn an engine-filled ShowSpec into deliverables.

Reuses the existing, verified tools — it does NOT re-derive the .ses byte format:
  1. render the locked FOH Channel Processing .md
  2. run the venue patcher (apply_show_TEMPLATE*.py) per the send-it procedure
  3. byte-verify (patcher PASS + exact output size)
  4. packet PDF (show-packet-builder-template.py), channel-processing HTML/PDF,
     input-list .xlsx
  5. stage a wiki bundle

Only memo + fsq have the calibrated .ses pipeline; other venues get paperwork
only. Crowd-rig channels (ch=None) are documented in the packet but never
written to the .ses.
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent


# ────────────────────────────────────────────────────────────────────────────
# 1. locked MD
# ────────────────────────────────────────────────────────────────────────────
def _md_channels(spec):
    """Channels that go to the .ses: numbered, non-crowd, non-spare, with a mic."""
    out = []
    for c in spec.channels:
        if c.is_crowd or c.ch is None:
            continue
        if (c.name or "").upper() == "SPARE" or not (c.mic or "").strip():
            continue
        out.append(c)
    return sorted(out, key=lambda c: c.ch)


def render_md(kn, spec) -> str:
    venue = kn.venue(spec.venue) or {}
    console = venue.get("console_label", "DiGiCo Quantum 225")
    lines = [
        f"# {spec.show_name} — FOH Channel Processing",
        f"## {venue.get('name', spec.venue)} · {console} · {spec.show_date}",
        "*Active channels only. Band order: B4 (high) → B3 → B2 → B1 (low).*",
        "",
    ]
    if spec.artist_profile:
        lines += [f"### Artist — performance sound ({spec.artist or 'artist'})",
                  spec.artist_profile, ""]
    for c in _md_channels(spec):
        lines.append(f"## Ch {c.ch} | {c.name} | {c.mic}")
        hpf = int(round(c.hpf)) if c.hpf else 20
        lpf = "OFF" if (c.lpf is None or c.lpf >= 20000) else int(round(c.lpf))
        lines.append(f"HPF: {hpf} | LPF: {lpf}")
        by_b = {b.b: b for b in c.bands}
        for bnum in (4, 3, 2, 1):
            b = by_b.get(bnum)
            lines.append(b.md_line() if b else f"B{bnum}: FLAT")
        if getattr(c, "research", ""):
            lines.append(f"*Research: {c.research}*")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


# ────────────────────────────────────────────────────────────────────────────
# 2-3. patcher + verify
# ────────────────────────────────────────────────────────────────────────────
def run_patcher(kn, spec, dest_ses: Path, md_path: Path):
    """Run the venue patcher. Returns (ok, log)."""
    venue = kn.venue(spec.venue) or {}
    if not venue.get("pipeline"):
        return False, f"Venue '{spec.venue}' has no calibrated .ses pipeline."
    template = kn.template_path(spec.venue)
    patcher = kn.patcher_path(spec.venue)
    if not template or not template.exists():
        return False, f"Template not found: {template}"
    if not patcher or not patcher.exists():
        return False, f"Patcher not found: {patcher}"

    cmd = [sys.executable, str(patcher), "--src", str(template),
           "--dest", str(dest_ses), "--md", str(md_path)]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    log = proc.stdout + ("\n" + proc.stderr if proc.stderr else "")

    ok = proc.returncode == 0 and "PASS" in proc.stdout and "FAIL" not in proc.stdout
    # exact-size gate
    want = venue.get("output_bytes")
    if dest_ses.exists() and want:
        got = dest_ses.stat().st_size
        if got != want:
            ok = False
            log += f"\nSIZE FAIL: {got} != {want}"
        else:
            log += f"\nSIZE OK: {got} bytes"
    elif not dest_ses.exists():
        ok = False
        log += "\nNO OUTPUT FILE"
    return ok, log


_PATCHER_CACHE = {}


def _import_patcher(kn, venue):
    """Import the venue's patcher module so the readback reuses ITS real byte
    offsets/helpers. The patcher carries the offset tripwire and gets
    recalibrated when a template is resaved; a private copy of the constants
    here would silently drift (it did — the FSQ recalibration of 2026-06-21
    moved SURF_BASE/SCAN and this readback kept the old values, false-failing
    every FSQ build). Importing keeps the two in lockstep forever."""
    if venue in _PATCHER_CACHE:
        return _PATCHER_CACHE[venue]
    import importlib.util
    p = kn.patcher_path(venue)
    if not p or not p.exists():
        _PATCHER_CACHE[venue] = None
        return None
    mspec = importlib.util.spec_from_file_location(f"_patcher_{venue}", p)
    mod = importlib.util.module_from_spec(mspec)
    mspec.loader.exec_module(mod)        # both patchers gate __main__, safe to import
    _PATCHER_CACHE[venue] = mod
    return mod


def readback_check(kn, spec, dest_ses: Path):
    """Independent HPF readback of one channel against the MD, using the venue
    patcher's own offsets (the send-it verify gate). Returns (ok, notes)."""
    data = open(dest_ses, "rb").read()
    chans = _md_channels(spec)
    if not chans:
        return True, "no channels to read back"
    # prefer a channel with both an HPF and a shelf band, else any HPF, else first
    target = (next((c for c in chans
                    if c.hpf and any(b.type == "SHELF" for b in c.bands)), None)
              or next((c for c in chans if c.hpf), None)
              or chans[0])
    mod = _import_patcher(kn, spec.venue)
    if mod is None:
        return False, "patcher not importable for readback"
    try:
        if spec.venue == "fsq":
            return _readback_fsq(mod, data, target)
        return _readback_memo(mod, data, target)
    except Exception as e:  # noqa
        return False, f"readback error: {e}"


def _readback_fsq(mod, data, ch):
    import struct
    bb = mod._block_bounds(data, ch.ch)
    if bb is None:
        return False, f"Ch{ch.ch} '{ch.name}': block not found / name not unique"
    lo, hi = bb
    lpf_vo = mod._lpf_value_offset(data, lo, hi)
    if lpf_vo is None:
        return False, f"Ch{ch.ch} '{ch.name}': LPF record not found"
    hpf_stored = struct.unpack_from("<f", data, lpf_vo + 0x10)[0]   # HPF rides LPF+0x10
    want = (int(round(ch.hpf)) if ch.hpf else 20) * mod.HPF_SCALE
    ok = abs(hpf_stored - want) < 1.0
    return ok, (f"Ch{ch.ch} '{ch.name}' HPF stored {hpf_stored:.1f} "
                f"want {want:.1f} {'OK' if ok else 'MISMATCH'}")


def _readback_memo(mod, data, ch):
    import struct
    start, _size = mod.strip_region(ch.ch)
    hpf_stored = struct.unpack_from("<f", data, start + mod.HPF_REL)[0]
    want = (int(round(ch.hpf)) if ch.hpf else 20) * mod.HPF_SCALE
    ok = abs(hpf_stored - want) < 1.0
    return ok, (f"Ch{ch.ch} '{ch.name}' HPF stored {hpf_stored:.1f} "
                f"want {want:.1f} {'OK' if ok else 'MISMATCH'}")


# ────────────────────────────────────────────────────────────────────────────
# 4. paperwork — packet PDF, input list xlsx, channel-processing HTML, review
# ────────────────────────────────────────────────────────────────────────────
def _load_packet_builder(kn):
    import importlib.util
    p = kn.audio_root / "show-packet-builder-template.py"
    spec = importlib.util.spec_from_file_location("packet_builder", p)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _fmt_freq(hz):
    if hz is None:
        return "—"
    return f"{hz/1000:g} kHz" if hz >= 1000 else f"{int(round(hz))} Hz"


def _patch_label(ch):
    return f"Local {ch.ch}" if ch.ch is not None else "—"


def _eq_bands_for_pdf(c):
    """Build the packet builder's band-row list: HPF, B4..B1 (console order), LPF."""
    rows = [{"band": "LC", "freq": _fmt_freq(c.hpf) if c.hpf else "—",
             "type": "LC" if c.hpf else "OFF", "gain": "—",
             "q": "18 dB/oct" if c.hpf else "—", "notes": "HPF"}]
    by_b = {b.b: b for b in c.bands}
    label = {4: "4", 3: "3", 2: "2", 1: "1"}
    for bnum in (4, 3, 2, 1):
        b = by_b.get(bnum)
        if not b:
            rows.append({"band": label[bnum], "freq": "—", "type": "OFF",
                         "gain": "—", "q": "—", "notes": "—"})
            continue
        note = b.role if hasattr(b, "role") else ""
        deq = ""
        if b.deq:
            deq = (f"DYNAMIC: Thresh {int(round(b.deq['thr']))}dBFS / "
                   f"Att {int(round(b.deq['atk_ms']))}ms / "
                   f"Rel {int(round(b.deq['rel_ms']))}ms")
        rows.append({"band": label[bnum], "freq": _fmt_freq(b.freq),
                     "type": "Shelf" if b.type == "SHELF" else "Bell",
                     "gain": f"{int(round(b.gain)):+d} dB", "q": f"{b.q:g}",
                     "notes": deq or ""})
    lpf_on = c.lpf is not None and c.lpf < 20000
    rows.append({"band": "HC", "freq": _fmt_freq(c.lpf) if lpf_on else "—",
                 "type": "HC" if lpf_on else "OFF", "gain": "—",
                 "q": "12 dB/oct" if lpf_on else "—", "notes": "LPF"})
    return rows


def build_packet_pdf(kn, spec, out_path):
    mod = _load_packet_builder(kn)
    accent = {"DRUMS": mod.DRUMS_BAR, "RHYTHM": mod.GUITAR_BAR, "PIANO": mod.PIANO_BAR,
              "STRINGS": mod.STRINGS_BAR, "HORNS": mod.HORNS_BAR, "VOCALS": mod.VOCALS_BAR,
              "AMBIENT": mod.AMBIENT_BAR, "SPARE": mod.TBD_BAR}
    venue = kn.venue(spec.venue) or {}

    il, eqc = [], []
    for c in spec.channels:
        il.append({
            "ch": "" if c.ch is None else c.ch,
            "name": c.name or "SPARE",
            "mic": c.mic or "—",
            "section": c.section or "SPARE",
            "patch": _patch_label(c),
            "phantom_48v": bool(c.phantom) and not c.ribbon,
            "stand": c.stand or "—",
            "notes": ("⚑ NO 48V (ribbon) " if c.ribbon else "") + (c.notes or ""),
        })
        if (c.name or "").upper() == "SPARE":
            continue
        summary = c.eq_summary or ""
        if c.comp:
            summary += (f"  COMP: Thr {c.comp['thr']}dB / {c.comp['ratio']}:1 / "
                        f"Att {c.comp['atk_ms']}ms / Rel {c.comp['rel_ms']}ms.")
        if c.gate:
            summary += "  GATE: on (dial threshold/release at soundcheck)."
        summary += "  (Comp/gate are set by hand — not written to the .ses.)"
        eqc.append({
            "ch": "" if c.ch is None else c.ch, "name": c.name, "mic": c.mic or "—",
            "section": c.section or "SPARE",
            "accent": accent.get(c.section, mod.TBD_BAR),
            "mic_notes": c.mic_notes or "—",
            "bands": _eq_bands_for_pdf(c), "summary": summary or "—",
        })

    show_data = {
        "show_name": spec.show_name, "venue": venue.get("name", spec.venue),
        "date": spec.show_date, "console_label": venue.get("console_label", "DiGiCo Quantum 225"),
        "foh_engineer": spec.foh_engineer, "mon_engineer": spec.mon_engineer,
        "show_time": spec.show_time, "rev": spec.rev,
        "channel_count": len([c for c in spec.channels if (c.name or '').upper() != 'SPARE']),
        "style_note": f"{spec.artist or '—'} · {spec.genre or '—'}",
    }
    mod.build_show_packet(str(out_path), show_data, il, eqc, console="digico")
    return out_path


def build_input_list_xlsx(kn, spec, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    venue = kn.venue(spec.venue) or {}
    ws["A1"] = spec.show_name
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = (f"{venue.get('name', spec.venue)} · {venue.get('console_label','')} · "
                f"{spec.show_date} · {spec.rev}")
    hdrs = ["Ch", "Instrument", "Mic / DI", "Patch", "48V", "Stand", "Notes"]
    ws.append([])
    ws.append(hdrs)
    hr = ws.max_row
    for cidx in range(1, len(hdrs) + 1):
        cell = ws.cell(row=hr, column=cidx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.alignment = Alignment(horizontal="center")
    for c in spec.channels:
        ws.append([
            "" if c.ch is None else c.ch, c.name or "SPARE", c.mic or "—",
            _patch_label(c), "✓" if (c.phantom and not c.ribbon) else "",
            c.stand or "—", ("NO 48V (ribbon) " if c.ribbon else "") + (c.notes or ""),
        ])
    widths = [6, 22, 26, 12, 6, 10, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(out_path)
    return out_path


def render_html(kn, spec, md_text, out_path):
    venue = kn.venue(spec.venue) or {}
    rows = ""
    import html
    for line in md_text.splitlines():
        if line.startswith("## Ch"):
            rows += f"<h3>{html.escape(line[3:])}</h3>"
        elif line.startswith("#"):
            continue
        elif line.startswith("B") or line.startswith("HPF"):
            rows += f"<div class='band'>{html.escape(line)}</div>"
        elif line.startswith("*"):
            rows += f"<p class='muted'>{html.escape(line.strip('*'))}</p>"
    rev_html = "".join(f"<li>{html.escape(r.line())}</li>" for r in spec.reverbs)
    doc = f"""<!doctype html><html><head><meta charset='utf-8'>
<title>{html.escape(spec.show_name)} — FOH Channel Processing</title>
<style>body{{font-family:-apple-system,Arial,sans-serif;max-width:900px;margin:2rem auto;color:#1a1a2e}}
h1{{color:#1A3A5C}} h3{{background:#0F3460;color:#fff;padding:.3rem .6rem;border-radius:4px;margin-top:1.2rem}}
.band{{font-family:Consolas,monospace;padding:.1rem .6rem}} .muted{{color:#666}}
.profile{{background:#F4F0E8;border-left:4px solid #E94560;padding:.6rem .9rem;margin:1rem 0;line-height:1.4}}
.rev li{{font-family:Consolas,monospace;font-size:.85rem;margin:.3rem 0}}</style></head><body>
<h1>{html.escape(spec.show_name)}</h1>
<p><b>{html.escape(venue.get('name', spec.venue))}</b> · {html.escape(venue.get('console_label',''))} ·
{html.escape(spec.show_date)} · {html.escape(spec.artist)} · {html.escape(spec.genre)}</p>
{f"<div class='profile'><b>Artist — performance sound:</b> {html.escape(spec.artist_profile)}</div>" if spec.artist_profile else ""}
{rows}
<h2>Seventh Heaven Pro — Reverb</h2><p class='muted'>{html.escape(spec.reverb_note)}</p>
<ul class='rev'>{rev_html}</ul>
</body></html>"""
    Path(out_path).write_text(doc, encoding="utf-8")
    return out_path


def build_review_pdf(kn, spec, out_path):
    """The 'quick list for approval' as a one/two-page PDF."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    venue = kn.venue(spec.venue) or {}
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1A3A5C"))
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#374151"))
    sm = ParagraphStyle("sm", parent=styles["Normal"], fontSize=8, leading=10)
    story = [Paragraph(f"{spec.show_name} — Review", h),
             Paragraph(f"{venue.get('name', spec.venue)} · {venue.get('console_label','')} · "
                       f"{spec.show_date} · {spec.artist} · {spec.genre}", sub),
             Paragraph(f"FOH {spec.foh_engineer} · MON {spec.mon_engineer} · {spec.show_time} · {spec.rev} · "
                       f"EQ {'ON' if spec.eq_on else 'off'} · Comp {'ON' if spec.comp_on else 'off'}", sub),
             Spacer(1, 10)]
    data = [["Ch", "Instrument", "Mic / DI", "48V", "Bands", "Notes"]]
    for c in spec.channels:
        bands = "FLAT" if not c.bands else " ".join(
            f"B{b.b}:{int(round(b.gain)):+d}@{int(round(b.freq))}" for b in sorted(c.bands, key=lambda x: x.b))
        data.append([
            "" if c.ch is None else str(c.ch),
            Paragraph(c.name or "SPARE", sm),
            Paragraph(c.mic or "—", sm),
            "✓" if (c.phantom and not c.ribbon) else ("NO" if c.ribbon else ""),
            Paragraph(bands, sm),
            Paragraph(("RIBBON " if c.ribbon else "") + (c.notes or ""), sm),
        ])
    t = Table(data, colWidths=[0.4*inch, 1.3*inch, 1.5*inch, 0.4*inch, 2.4*inch, 1.3*inch], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8), ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9CA3AF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    story += [t, Spacer(1, 12), Paragraph("Seventh Heaven Pro — Reverb", sub),
              Paragraph(spec.reverb_note, sm), Spacer(1, 4)]
    for r in spec.reverbs:
        story.append(Paragraph(r.line(), sm))
    SimpleDocTemplate(str(out_path), pagesize=letter,
                      leftMargin=0.5*inch, rightMargin=0.5*inch,
                      topMargin=0.5*inch, bottomMargin=0.5*inch).build(story)
    return out_path


# ────────────────────────────────────────────────────────────────────────────
# 5. orchestrate
# ────────────────────────────────────────────────────────────────────────────
def build_all(kn, spec, *, write_ses=True):
    """Produce all deliverables in the show folder. Returns a result dict."""
    venue = kn.venue(spec.venue) or {}
    folder = kn.show_folder(spec.venue, spec.folder_name())
    if folder is None:
        raise RuntimeError("audio_root not configured — cannot resolve show folder.")
    folder.mkdir(parents=True, exist_ok=True)
    base = folder / spec.show_name
    result = {"folder": str(folder), "files": {}, "ses_ok": None,
              "ses_log": "", "readback": "", "warnings": []}

    # package (source of truth)
    spec_path = folder / f"{spec.slug()}.spec.json"
    spec.save(spec_path)
    result["files"]["spec"] = str(spec_path)

    # locked MD + HTML
    md_text = render_md(kn, spec)
    md_path = folder / f"{spec.show_name} - FOH Channel Processing.md"
    md_path.write_text(md_text, encoding="utf-8")
    result["files"]["md"] = str(md_path)
    html_path = folder / f"{spec.show_name} - FOH Channel Processing.html"
    render_html(kn, spec, md_text, html_path)
    result["files"]["html"] = str(html_path)

    # .ses (memo/fsq only)
    if write_ses and venue.get("pipeline"):
        dest = folder / f"{spec.show_name}.ses"
        ok, log = run_patcher(kn, spec, dest, md_path)
        result["ses_ok"], result["ses_log"] = ok, log
        if ok:
            rok, notes = readback_check(kn, spec, dest)
            result["readback"] = notes
            result["ses_ok"] = ok and rok
            result["files"]["ses"] = str(dest)
            if not rok:
                result["warnings"].append("Readback mismatch — inspect before loading.")
        else:
            result["warnings"].append("Patcher did not PASS — .ses not trustworthy.")
    elif not venue.get("pipeline"):
        result["warnings"].append(f"{venue.get('name', spec.venue)} has no .ses pipeline — paperwork only.")

    # packet PDF + xlsx + review PDF
    try:
        result["files"]["packet"] = str(build_packet_pdf(kn, spec, base.with_name(f"{spec.show_name} - Show Packet.pdf")))
    except Exception as e:  # noqa
        result["warnings"].append(f"Packet PDF failed: {e}")
    try:
        result["files"]["xlsx"] = str(build_input_list_xlsx(kn, spec, folder / f"{spec.show_name} - Input List.xlsx"))
    except Exception as e:  # noqa
        result["warnings"].append(f"Input list xlsx failed: {e}")
    try:
        result["files"]["review"] = str(build_review_pdf(kn, spec, folder / f"{spec.show_name} - Review.pdf"))
    except Exception as e:  # noqa
        result["warnings"].append(f"Review PDF failed: {e}")

    return result


def wiki_bundle(result):
    """List the files that should go to the wiki (push runs via the
    wiki-publish / fsq-wiki-push skill on the Mac — it has LAN/SSH/git)."""
    keep = ("md", "html", "ses", "packet", "xlsx", "review")
    return [result["files"][k] for k in keep if k in result["files"]]

