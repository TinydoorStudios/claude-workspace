#!/usr/bin/env python3
"""Mac side of staff-booking seeding: append new bookings as rows in the sheet.

Reads the bookings JSON (from seed_bookings.py --json), opens the local
advance-list.xlsx, and appends any booking that isn't already a row (matched on
event + date + venue + artist). Prints the comma-separated ids it handled to
STDOUT so generate.command can stamp them seeded; everything else goes to STDERR.

  python3 append_bookings.py --list advance-list.xlsx --data bookings.json
"""
import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import fieldspec as fs

from openpyxl import load_workbook

INPUT_SHEET = "Advance List"


def _s(v):
    return "" if v is None else str(v).strip()


def norm(v):
    return re.sub(r"\s+", " ", _s(v)).lower()


def find_header_row(ws):
    labels = {lbl.lower() for (lbl, _k, _c) in fs.ALL_COLUMNS}
    best_row, best = 2, 0
    for r in range(1, 8):
        hits = sum(1 for c in ws[r] if _s(c.value).lower() in labels)
        if hits > best:
            best, best_row = hits, r
    return best_row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", required=True, type=Path)
    ap.add_argument("--data", required=True, type=Path)
    args = ap.parse_args()

    bookings = json.loads(args.data.read_text()) if args.data.exists() else []
    if not bookings:
        return
    wb = load_workbook(args.list)
    ws = wb[INPUT_SHEET] if INPUT_SHEET in wb.sheetnames else wb.worksheets[0]

    hrow = find_header_row(ws)
    # key -> column, from the header row
    key_col = {}
    for c in range(1, ws.max_column + 1):
        lbl = _s(ws.cell(hrow, c).value)
        if lbl in fs.LABEL_TO_KEY:
            key_col[fs.LABEL_TO_KEY[lbl]] = c
    c_artist = key_col.get("artist_name")
    c_venue = key_col.get("venue")
    c_date = key_col.get("event_date")
    if not c_artist:
        print("no Artist Name column; cannot seed", file=sys.stderr)
        return

    # existing (artist, venue, date) identities already in the sheet
    existing = set()
    last = hrow
    for r in range(hrow + 1, ws.max_row + 1):
        a = _s(ws.cell(r, c_artist).value)
        if a:
            last = r
            existing.add((norm(a),
                          norm(ws.cell(r, c_venue).value) if c_venue else "",
                          _s(ws.cell(r, c_date).value)[:10] if c_date else ""))

    handled, appended = [], 0
    row = last + 1
    for b in bookings:
        handled.append(str(b["id"]))
        ident = (norm(b.get("artist_name")), norm(b.get("venue")),
                 _s(b.get("event_date"))[:10])
        if ident in existing:
            continue  # already represented (staff typed it, or a prior seed)
        for key, col in key_col.items():
            val = b.get(key)
            if val not in (None, ""):
                ws.cell(row, col).value = _s(val)
        existing.add(ident)
        row += 1
        appended += 1

    if appended:
        wb.save(args.list)
    print(f"appended {appended} new booking row(s)", file=sys.stderr)
    # STDOUT: ids to stamp seeded (handled whether appended or already present)
    print(",".join(handled))


if __name__ == "__main__":
    main()
