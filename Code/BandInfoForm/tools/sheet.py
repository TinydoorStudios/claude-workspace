#!/usr/bin/env python3
"""Reader for the 3CDC Advance List spreadsheet (tools/lists/advance_list_template.xlsx).

Normalizes each data row to a dict. Skips the gray example rows and blanks.
Columns: Event Name, Event Date, Venue, Series, Slot, Set Time, Artist Name,
Contact Email, Notes.
"""
import datetime as dt
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import fieldspec as fs

# column label (lowercased) -> internal key, for every column in the spec
FIELDS = {lbl.lower(): key for (lbl, key, _ch) in fs.ALL_COLUMNS}
FIELDS["notes"] = "notes"  # tolerate a legacy Notes column if present


def _norm_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s  # leave as-is; downstream parser will try again


def _band_owned(wb):
    """Cells the band filled (from merge_status' hidden map) -> their value.

    A band-filled cell is a live mirror of the form, NOT a Brian override, so the
    importer must read it as blank; only cells Brian typed himself win."""
    meta = {}
    if "_advance_meta" in wb.sheetnames:
        for row in wb["_advance_meta"].iter_rows(values_only=True):
            if row and row[0]:
                meta[str(row[0])] = "" if len(row) < 2 or row[1] is None else str(row[1])
    return meta


def read_advance_sheet(path):
    from openpyxl import load_workbook
    wb = load_workbook(Path(path), data_only=True)
    owned = _band_owned(wb)
    ws = wb["Advance List"] if "Advance List" in wb.sheetnames else wb.active
    # Find the header row wherever it is (a group-label row may sit above it) and
    # map columns by LABEL, not position — so columns can be reordered or removed.
    header, header_row, best = {}, 1, 0
    for ridx in range(1, 8):
        m = {}
        for cell in ws[ridx]:
            key = FIELDS.get(str(cell.value or "").strip().lower())
            if key:
                m[cell.column] = key
        if len(m) > best:
            best, header_row, header = len(m), ridx, m
    if not header:
        return []
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1):
        rec = {}
        for cell in row:
            key = header.get(cell.column)
            if key:
                addr = f"r{cell.row}c{cell.column}"
                if addr in owned and str(cell.value or "") == owned[addr]:
                    continue  # band-filled mirror, not a Brian override
                rec[key] = cell.value
        artist = (str(rec.get("artist_name") or "")).strip()
        notes = (str(rec.get("notes") or "")).strip().upper()
        if not artist or "EXAMPLE ROW" in notes or artist.upper().startswith("EXAMPLE"):
            continue
        rec = {k: (str(v).strip() if isinstance(v, str) else v) for k, v in rec.items()}
        rec["artist_name"] = artist
        rec["event_date"] = _norm_date(rec.get("event_date"))
        rows.append(rec)
    return rows
