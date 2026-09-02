#!/usr/bin/env python3
"""Generate advance_status.xlsx — the companion view.

One row per band/show: your sheet input + what the band submitted on the form +
the advance state (emailed / responded / follow-up due). This is the read-back you
watch; your input spreadsheet stays yours to edit. Refreshed by generate.command.
"""
import datetime as dt
import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
for _cand in (HERE.parent, HERE.parent / "app"):
    if (_cand / "advance_db.py").exists():
        sys.path.insert(0, str(_cand))
        break
sys.path.insert(0, str(HERE))
import advance_db as db
import fieldspec as fs

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_JSON = "--json" in sys.argv[1:]
_args = [a for a in sys.argv[1:] if not a.startswith("--")]
OUT = Path(_args[0]) if _args else (HERE / "advance_status.xlsx")
NAVY = "1A3A5C"
STATE_FILL = {
    "queued": "E5E7EB", "awaiting": "FEF3C7", "followup_due": "FFE4B5",
    "followup_sent": "DBEAFE", "responded": "DCFCE7",
}


def yn(v):
    return "Yes" if v is True else ("No" if v is False else "")


def g(sub, key):
    if not sub:
        return ""
    v = sub.get(key)
    if v in (None, ""):
        v = (sub.get("data") or {}).get(key, "")
    return "" if v is None else v


def d(ts):
    return ts.strftime("%Y-%m-%d") if ts else ""


FOLLOWUP_DAYS = 10  # matches the advance_status view + n8n check


def followup_due(st):
    """Date a follow-up is due — only while the advance isn't completed."""
    if st["state"] == "responded":          # completed box is checked
        return ""
    if st.get("followup_sent_at"):
        return "sent " + d(st["followup_sent_at"])
    if st.get("email_sent_at"):
        return (st["email_sent_at"] + dt.timedelta(days=FOLLOWUP_DAYS)).strftime("%Y-%m-%d")
    return ""


# (header, width, extractor(st, sub, act))
COLS = [
    ("Band", 30, lambda st, sub, a: st["band"]),
    ("Event", 24, lambda st, sub, a: (a or {}).get("event_name", "")),
    ("Venue", 16, lambda st, sub, a: st["venue"]),
    ("Date", 12, lambda st, sub, a: d(st["show_date"]) if st["show_date"] else ""),
    ("Slot", 14, lambda st, sub, a: (a or {}).get("slot", "")),
    ("Set Length", 13, lambda st, sub, a: (a or {}).get("set_time", "")),
    ("Status", 14, lambda st, sub, a: st["state"]),
    ("Email Sent", 12, lambda st, sub, a: d(st["email_sent_at"])),
    ("Follow-up Due", 13, lambda st, sub, a: followup_due(st)),
    ("Completed", 11, lambda st, sub, a: "Yes" if st["state"] == "responded" else ""),
    ("Responded", 12, lambda st, sub, a: d(st["responded_at"])),
    ("Contact Name", 18, lambda st, sub, a: g(sub, "contact_name")),
    ("Contact Email", 22, lambda st, sub, a: g(sub, "contact_email") or (st.get("contact_email") or "")),
    ("Contact Phone", 15, lambda st, sub, a: g(sub, "contact_phone")),
    ("Performers", 11, lambda st, sub, a: g(sub, "performers")),
    ("Monitors", 10, lambda st, sub, a: g(sub, "monitors")),
    ("Own IEMs", 10, lambda st, sub, a: yn(sub.get("own_iems")) if sub else ""),
    ("Split", 8, lambda st, sub, a: g(sub, "split_snake")),
    ("Stage", 13, lambda st, sub, a: g(sub, "stage_type")),
    ("Own Engineer", 22, lambda st, sub, a: g(sub, "own_engineer")),
    ("Merch", 8, lambda st, sub, a: yn(sub.get("merch")) if sub else ""),
    ("Band Tent", 22, lambda st, sub, a: g(sub, "band_tent")),
    ("Large Vehicle", 12, lambda st, sub, a: yn(sub.get("large_vehicle")) if sub else ""),
    ("Stage Plot", 30, lambda st, sub, a: g(sub, "stage_plot_desc")),
    ("Backline", 24, lambda st, sub, a: g(sub, "backline")),
    ("Scenic", 22, lambda st, sub, a: g(sub, "scenic")),
    ("Lighting", 20, lambda st, sub, a: g(sub, "lighting")),
    ("Changed Notes", 26, lambda st, sub, a: g(sub, "changed_notes")),
    ("Additional", 26, lambda st, sub, a: g(sub, "additional")),
]


def act_for(cur, st):
    """The event act matching this show (artist + venue + date), for event/slot/set."""
    cur.execute(
        """SELECT ea.slot, ea.set_time, e.name AS event_name
           FROM event_acts ea JOIN events e ON e.id = ea.event_id
           WHERE ea.artist_id = %s
             AND COALESCE(e.venue,'') = COALESCE(%s,'')
             AND e.event_date IS NOT DISTINCT FROM %s
           LIMIT 1""",
        (st["artist_id"], st["venue"], st["show_date"]),
    )
    return cur.fetchone() or {}


# internal-key -> how to pull that band's FORM answer for the fill-the-blanks merge
def _band_fields(sub):
    if not sub:
        return {}
    out = {
        "contact_email": g(sub, "contact_email"),
        "contact_name": g(sub, "contact_name"),
        "contact_phone": g(sub, "contact_phone"),
        "stage_type": g(sub, "stage_type"),
        "monitors": g(sub, "monitors"),
        "own_iems": yn(sub.get("own_iems")),
        "split_snake": g(sub, "split_snake"),
        "stage_plot_desc": g(sub, "stage_plot_desc"),
        "input_notes": g(sub, "input_notes"),
        "backline": g(sub, "backline"),
        "own_engineer": g(sub, "own_engineer"),
        "scenic": g(sub, "scenic"),
        "lighting": g(sub, "lighting"),
        "merch": yn(sub.get("merch")),
        "band_tent": g(sub, "band_tent"),
        "performers": g(sub, "performers"),
        "large_vehicle": yn(sub.get("large_vehicle")),
    }
    return {k: ("" if v is None else str(v)) for k, v in out.items() if v not in (None, "")}


def records():
    """Structured per-advance data for the Mac-side merge into advance-list.xlsx."""
    out = []
    with db.get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM advance_status ORDER BY show_date NULLS LAST, band")
        for st in cur.fetchall():
            sub = db.newest_submission(cur, st["artist_id"])
            act = act_for(cur, st)
            out.append({
                "band": st["band"],
                "venue": st["venue"] or "",
                "date": st["show_date"].isoformat() if st["show_date"] else "",
                "slot": (act or {}).get("slot", ""),
                "state": st["state"],
                "email_sent": d(st["email_sent_at"]),
                "followup_due": followup_due(st),
                "completed": "Yes" if st["state"] == "responded" else "",
                "responded": d(st["responded_at"]),
                "changed_notes": g(sub, "changed_notes"),
                "additional": g(sub, "additional"),
                "band_fields": _band_fields(sub),
            })
    return out


def emit_json(path):
    path = Path(path)
    recs = records()
    path.write_text(json.dumps(recs, indent=2, default=str))
    print(f"Wrote {path} — {len(recs)} advance(s)")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Advance Status"
    thin = Side(style="thin", color="D9DEE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (name, width, _fn) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    n = 0
    with db.get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM advance_status ORDER BY show_date NULLS LAST, band")
        statuses = cur.fetchall()
        for st in statuses:
            sub = db.newest_submission(cur, st["artist_id"])
            act = act_for(cur, st)
            row = n + 2
            for i, (_name, _w, fn) in enumerate(COLS, start=1):
                cell = ws.cell(row=row, column=i, value=fn(st, sub, act))
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            # tint the Status cell by state
            fill = STATE_FILL.get(st["state"])
            if fill:
                ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=fill)
            n += 1

    wb.save(OUT)
    print(f"Wrote {OUT} — {n} advance(s)")


if __name__ == "__main__":
    if _JSON:
        emit_json(OUT)
    else:
        main()
