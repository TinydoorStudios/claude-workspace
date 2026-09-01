#!/usr/bin/env python3
"""Build tools/lists/advance_list_template.xlsx from fieldspec. Reproducible."""
from pathlib import Path
import sys

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import fieldspec as fs

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = HERE / "lists" / "advance_list_template.xlsx"

NAVY, ACCENT, LIGHT, GRAY = "1A3A5C", "2E6DA4", "EAF1F8", "9AA5B1"
GROUP_FILLS = ["24506E", "2E6DA4", "3E7CB1"]
WIDE = {"Stage Plot", "Backline", "Scenic", "Input Notes", "Lighting", "Notes",
        "Event Name", "Artist Name", "Contact Email"}

wb = Workbook()
wb.remove(wb.active)  # drop the default "Sheet"

# hidden lists sheet for dropdowns (handles choices containing commas)
wl = wb.create_sheet("_lists")
ranges = {}
col = 1
for lbl, key, choices in fs.ALL_COLUMNS:
    if not choices:
        continue
    letter = get_column_letter(col)
    wl.cell(row=1, column=col, value=key)
    for i, ch in enumerate(choices, start=2):
        wl.cell(row=i, column=col, value=ch)
    ranges[key] = f"_lists!${letter}$2:${letter}${len(choices) + 1}"
    col += 1
wl.sheet_state = "hidden"

ws = wb.create_sheet("Advance List", 0)
thin = Side(style="thin", color="D9DEE5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cols = fs.ALL_COLUMNS
# row 1: group headers (merged)
ci = 1
for gi, (label, span) in enumerate(fs.GROUPS):
    ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + span - 1)
    c = ws.cell(row=1, column=ci, value=label)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=GROUP_FILLS[gi % 3])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ci += span
# row 2: column labels
for i, (lbl, key, choices) in enumerate(cols, start=1):
    c = ws.cell(row=2, column=i, value=lbl)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = 26 if lbl in WIDE else 14
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 30
ws.freeze_panes = "A3"

# example rows (a 2-act event) — gray italic, overwrite/delete
ex = {
    "event_name": "513 Airwaves w/ Inhailer", "event_date": "2026-09-20",
    "venue": "Fountain Square", "event_type": "Internal", "paying_band": "Yes",
    "mc": "Inhailer Radio", "dj": "n/a", "lead_name": "Lily", "lead_phone": "513-555-0000",
}
examples = [
    {**ex, "slot": "headliner", "set_time": "9:00p-10:00p",
     "artist_name": "Buffalo Wabs and the Price Hill Hustle",
     "contact_email": "booking@example.com"},
    {**ex, "slot": "opener", "set_time": "7:00p-7:45p",
     "artist_name": "The Cincy Suns", "contact_email": "manager@example.com"},
]
for r, row in enumerate(examples, start=3):
    for i, (lbl, key, ch) in enumerate(cols, start=1):
        c = ws.cell(row=r, column=i, value=row.get(key, ""))
        c.font = Font(italic=True, color=GRAY, size=10)
        c.border = border
    ws.cell(row=r, column=len(cols)).value = None  # (no Notes column now)

# blank rows + validations
LAST = 200
for r in range(5, 60):
    if r % 2 == 0:
        for i in range(1, len(cols) + 1):
            ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=LIGHT)
for r in range(3, LAST + 1):
    for i in range(1, len(cols) + 1):
        ws.cell(row=r, column=i).border = border

for i, (lbl, key, choices) in enumerate(cols, start=1):
    if not choices:
        continue
    dv = DataValidation(type="list", formula1=ranges[key], allow_blank=True)
    ws.add_data_validation(dv)
    letter = get_column_letter(i)
    dv.add(f"{letter}3:{letter}{LAST}")

# How to use
ws2 = wb.create_sheet("How to use")
ws2.column_dimensions["A"].width = 104
lines = [
    ("3CDC Advance List — the master source", True),
    ("", False),
    ("This sheet is the source of truth. It creates the advance email and the day-sheet", False),
    ("document. The band's advance FORM fills in whatever you leave blank in the band-detail", False),
    ("columns — you don't have to enter those yourself unless you want to override the band.", False),
    ("", False),
    ("One row per band, per event.", True),
    ("Rows sharing Event Name + Date + Venue are ONE event (a bill). Give each act its Slot.", False),
    ("Most events are 1-2 acts. Fill the EVENT columns once per event (any row in the group).", False),
    ("", False),
    ("The three column groups:", True),
    ("  EVENT — event name, date, venue, series, type, paying?, MC/DJ, site lead. You fill these.", False),
    ("  ACT — slot, set time, the exact artist name, and the email the advance goes to.", False),
    ("  BAND DETAILS — stage plot, monitors, scenic, merch, parking, contact, etc. Normally the", False),
    ("     band fills these via the form; type a value here only to override what the band submits.", False),
    ("", False),
    ("Merge rule: a value you put in this sheet WINS; the form fills anything you left blank.", True),
    ("", False),
    ("Artist Name is used verbatim on every email and document, and matched case-insensitively", False),
    ("to spot returning artists (their last submission pre-fills the form).", False),
    ("", False),
    ("Delete the gray example rows before using it for real.", False),
]
for r, (text, bold) in enumerate(lines, start=1):
    c = ws2.cell(row=r, column=1, value=text)
    c.font = Font(bold=bold, size=13 if r == 1 else 11, color=NAVY if bold else "1F2937")
    c.alignment = Alignment(wrap_text=True, vertical="center")

wb.save(OUT)
print("wrote", OUT, "-", len(cols), "columns")
