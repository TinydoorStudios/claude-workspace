#!/usr/bin/env python3
"""Fold the advance status + band form answers into advance-list.xlsx — one tab.

Runs on the Mac (openpyxl only, no database). The server hands over status.json
(one record per advanced band). This merges it into the live "Advance List" tab:

  • FILL THE BLANKS — wherever Brian left a band-detail cell empty, the band's form
    answer drops in, tinted blue so its source is obvious. His own typed cells are
    never overwritten (the spreadsheet wins; the form fills gaps).
  • STATUS BLOCK — appended to the right: Status (tinted by state), Advance
    Drafted, Follow-up Due, Completed, Responded, What Changed, Additional.

A hidden `_advance_meta` sheet records which cells the band filled, so a re-run
refreshes them to the current form answer, and a cell Brian later types over is
recognized as his and left alone. The reader (sheet.py) consults the same map so a
band-filled cell is never mistaken for a Brian override on the next generate.

  python3 merge_status.py --list advance-list.xlsx --data status.json
"""
import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from urllib.parse import quote

sys.path.insert(0, str(Path(__file__).resolve().parent))
import fieldspec as fs

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_SHEET = "Advance List"
META_SHEET = "_advance_meta"

NAVY = "1A3A5C"
BAND_TINT = "DDEBF7"          # band-provided answer
BANDING = "EAF1F8"           # the sheet's even-row banding
GROUP_FILL = "2E7D5B"        # STATUS group bar (green = auto)
STATE_FILL = {
    "queued": "E5E7EB", "awaiting": "FEF3C7", "followup_due": "FFE4B5",
    "followup_drafted": "DBEAFE", "responded": "C6EFCE",
}

# appended status columns: (label, json key, width)
STATUS_COLS = [
    ("Status", "state", 15), ("Advance Drafted", "advance_drafted", 14),
    ("Follow-up Due", "followup_due", 14), ("Completed", "completed", 11),
    ("Responded", "responded", 12), ("What Changed", "changed_notes", 26),
    ("Additional", "additional", 26),
]
FILL_KEYS = list(fs.BAND_KEYS) + ["contact_email"]
_thin = Side(style="thin", color="D9DEE5")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _s(v):
    return "" if v is None else str(v).strip()


def norm(v):
    return re.sub(r"\s+", " ", _s(v)).lower()


def norm_date(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
    s = _s(v)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s[:10] if s else ""


def banding(row):
    return (PatternFill("solid", fgColor=BANDING) if row % 2 == 0
            else PatternFill(fill_type=None))


def find_header_row(ws):
    """Row whose cells best match the spec labels (a group bar may sit above)."""
    labels = {lbl.lower() for (lbl, _k, _c) in fs.ALL_COLUMNS}
    best_row, best = 2, 0
    for r in range(1, 8):
        hits = sum(1 for c in ws[r] if _s(c.value).lower() in labels)
        if hits > best:
            best, best_row = hits, r
    return best_row


def load_meta(wb):
    meta = {}
    if META_SHEET in wb.sheetnames:
        ws = wb[META_SHEET]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                meta[str(row[0])] = "" if len(row) < 2 or row[1] is None else str(row[1])
    return meta


def save_meta(wb, meta):
    if META_SHEET in wb.sheetnames:
        del wb[META_SHEET]
    ws = wb.create_sheet(META_SHEET)
    ws.sheet_state = "veryHidden"
    for i, (addr, val) in enumerate(sorted(meta.items()), start=1):
        ws.cell(i, 1, addr)
        ws.cell(i, 2, val)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", required=True, type=Path)
    ap.add_argument("--data", required=True, type=Path)
    args = ap.parse_args()

    if not args.list.exists():
        raise SystemExit(f"no such workbook: {args.list}")
    records = json.loads(args.data.read_text()) if args.data.exists() else []
    index = {(norm(r["band"]), norm(r.get("venue")), norm_date(r.get("date"))): r
             for r in records}

    wb = load_workbook(args.list)
    if "Status" in wb.sheetnames:        # retire the old separate tab
        del wb["Status"]
    ws = wb[INPUT_SHEET] if INPUT_SHEET in wb.sheetnames else wb.worksheets[0]

    hrow = find_header_row(ws)
    data_start = hrow + 1

    # label -> column, from the actual header row (so columns can be reordered or
    # removed). The STATUS block is anchored past the last real input column, not
    # a hardcoded count.
    spec_labels = {lbl for (lbl, _k, _c) in fs.ALL_COLUMNS}
    col_of = {}
    for c in range(1, ws.max_column + 1):
        lbl = _s(ws.cell(hrow, c).value)
        if lbl in spec_labels:
            col_of[lbl] = c
    n_input = max(col_of.values()) if col_of else 0
    key_col = {fs.LABEL_TO_KEY[lbl]: c for lbl, c in col_of.items()
               if lbl in fs.LABEL_TO_KEY}
    fill_cols = [(k, key_col[k]) for k in FILL_KEYS if k in key_col]
    c_artist = key_col.get("artist_name")
    c_venue = key_col.get("venue")
    c_date = key_col.get("event_date")
    if not c_artist:
        raise SystemExit("couldn't locate the Artist Name column")

    meta_prev = load_meta(wb)
    meta_new = {}

    # last real data row
    last = data_start - 1
    for r in range(data_start, ws.max_row + 1):
        if _s(ws.cell(r, c_artist).value):
            last = r

    # ── fill the blanks, per data row ────────────────────────────────────────
    for r in range(data_start, last + 1):
        band = _s(ws.cell(r, c_artist).value)
        if not band or band.upper().startswith("EXAMPLE"):
            continue
        venue = _s(ws.cell(r, c_venue).value) if c_venue else ""
        date = norm_date(ws.cell(r, c_date).value) if c_date else ""
        rec = index.get((norm(band), norm(venue), date)) or {}
        bf = rec.get("band_fields", {})

        plot_rel = rec.get("stageplot_rel")
        for k, c in fill_cols:
            cell = ws.cell(r, c)
            addr = f"r{r}c{c}"
            # Stage Plot: when a file was filed, the cell becomes a clickable link
            # to it (relative to the workbook) instead of the typed description.
            link = plot_rel if k == "stage_plot_desc" and plot_rel else None
            B = link.rsplit("/", 1)[-1] if link else _s(bf.get(k, ""))
            prev = meta_prev.get(addr)
            managed = prev is not None
            owned = managed and _s(cell.value) == _s(prev)
            if B:
                if _s(cell.value) == "" or owned:
                    cell.value = B
                    cell.fill = PatternFill("solid", fgColor=BAND_TINT)
                    if link:
                        cell.hyperlink = quote(link, safe="/")
                        cell.font = Font(color="0563C1", underline="single")
                    meta_new[addr] = B
                elif managed:                 # he typed over it — it's his now
                    cell.fill = banding(r)
                    cell.hyperlink = None
            else:
                if owned:
                    cell.value = None
                    cell.fill = banding(r)
                    cell.hyperlink = None
                elif managed:
                    cell.fill = banding(r)

    # ── append the STATUS block ──────────────────────────────────────────────
    start = n_input + 1
    end = n_input + len(STATUS_COLS)
    for mr in [m for m in list(ws.merged_cells.ranges) if m.min_col >= start]:
        ws.unmerge_cells(str(mr))
    for r in range(1, ws.max_row + 1):       # clear any prior block
        for c in range(start, end + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()

    ws.merge_cells(start_row=hrow - 1, start_column=start, end_row=hrow - 1, end_column=end)
    gb = ws.cell(hrow - 1, start, "STATUS — auto, do not edit")
    gb.fill = PatternFill("solid", fgColor=GROUP_FILL)
    gb.font = Font(bold=True, color="FFFFFF")
    gb.alignment = Alignment(horizontal="center", vertical="center")
    for i, (lbl, _k, width) in enumerate(STATUS_COLS):
        c = start + i
        h = ws.cell(hrow, c, lbl)
        h.fill = PatternFill("solid", fgColor=NAVY)
        h.font = Font(bold=True, color="FFFFFF", size=10)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width

    for r in range(data_start, last + 1):
        band = _s(ws.cell(r, c_artist).value)
        if not band or band.upper().startswith("EXAMPLE"):
            continue
        venue = _s(ws.cell(r, c_venue).value) if c_venue else ""
        date = norm_date(ws.cell(r, c_date).value) if c_date else ""
        rec = index.get((norm(band), norm(venue), date)) or {}
        for i, (_lbl, k, _w) in enumerate(STATUS_COLS):
            c = start + i
            cell = ws.cell(r, c, rec.get(k, "") or "")
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = banding(r)
        state = rec.get("state")
        scell = ws.cell(r, start)
        scell.font = Font(bold=True)
        if state and STATE_FILL.get(state):
            scell.fill = PatternFill("solid", fgColor=STATE_FILL[state])

    save_meta(wb, meta_new)
    wb.save(args.list)
    print(f"Merged status into {args.list.name} — {len(records)} advance(s), "
          f"{len(meta_new)} band-filled cell(s).")


if __name__ == "__main__":
    main()
