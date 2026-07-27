"""Input List xlsx — column order, widths and colors per CLAUDE.md."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .render import CONSOLE_INFO_FIELDS, SECTION_COLORS, SECTION_ORDER

COLS = [
    ("Ch", 6),
    ("Instrument", 22),
    ("Mic/DI", 26),
    ("Split Patch", 12),
    ("48V", 6),
    ("Stand", 10),
    ("Notes", 32),
]
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(bottom=THIN)


def _fill(hexcolor: str) -> PatternFill:
    return PatternFill("solid", fgColor=hexcolor.lstrip("#"))


def _rig_info_tab(wb, sheet: dict, console: dict) -> None:
    """Location, console and I/O device blocks — skipped when nothing's filled in."""
    loc = sheet.get("location") or {}
    ci = sheet.get("console_info") or {}
    boxes = sheet.get("stageboxes") or []
    site = " · ".join(x for x in (loc.get("site"), loc.get("room")) if x)
    address = ", ".join(x for x in (loc.get("address"), loc.get("city"),
                                    " ".join(x for x in (loc.get("state"), loc.get("zip")) if x)) if x)
    loc_pairs = [(k, v) for k, v in [
        ("Project", loc.get("project")),
        ("Client / org", loc.get("client")),
        ("Site", site),
        ("Address", address),
    ] if v]
    con_pairs = [(label, ci.get(key)) for key, label in CONSOLE_INFO_FIELDS if ci.get(key) not in (None, "")]
    if not (loc_pairs or con_pairs or boxes):
        return

    ws = wb.create_sheet("Rig Info")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    row = 1

    def header(text: str, span: int = 2) -> None:
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = _fill("#111827")
        c.alignment = Alignment(vertical="center", indent=1)
        row += 1

    def pair(k, v) -> None:
        nonlocal row
        a = ws.cell(row=row, column=1, value=k)
        a.font = Font(name="Calibri", size=10, color="4B5563")
        a.border = BORDER
        b = ws.cell(row=row, column=2, value=v)
        b.font = Font(name="Calibri", size=10)
        b.border = BORDER
        row += 1

    if loc_pairs:
        header("Location")
        for k, v in loc_pairs:
            pair(k, v)
        row += 1
    if con_pairs:
        header("Console — " + console["label"])
        for k, v in con_pairs:
            pair(k, v)
        row += 1
    if boxes:
        header("I/O devices", span=4)
        for i, label in enumerate(["Device", "Notes", "Inputs", "Outputs"], start=1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = _fill("#E5E7EB")
        row += 1
        for b in boxes:
            meta = " · ".join(x for x in (b.get("location"), b.get("format"), b.get("notes")) if x)
            for i, val in enumerate([b.get("name"), meta, b.get("inputs"), b.get("outputs")], start=1):
                c = ws.cell(row=row, column=i, value=val if val not in (None, "") else "")
                c.font = Font(name="Calibri", size=10)
                c.border = BORDER
                if i in (3, 4):
                    c.alignment = Alignment(horizontal="center")
            row += 1


def build(sheet: dict, console: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Input List"

    for i, (label, width) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ncols = len(COLS)

    # Title bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=sheet.get("name", ""))
    c.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    c.fill = _fill(console["title_color"])
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    m = sheet.get("meta", {})
    bits = [
        sheet.get("venue_label"), console["label"], sheet.get("date"),
        f"Rev {sheet.get('rev')}",
        f"FOH {m.get('foh')}" if m.get("foh") else "",
        f"MON {m.get('mon')}" if m.get("mon") else "",
        m.get("showtime") or "",
    ]
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value="   ·   ".join(str(b) for b in bits if b))
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = _fill("#374151")
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Column headers
    for i, (label, _w) in enumerate(COLS, start=1):
        c = ws.cell(row=3, column=i, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = _fill("#111827")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    rows = sorted(sheet.get("inputs", []), key=lambda r: r.get("ch") or 0)
    grouped: dict[str, list[dict]] = {}
    for r in rows:
        if not (r.get("name") or r.get("mic") or r.get("instrument")):
            continue
        grouped.setdefault(r.get("section") or "SPARE", []).append(r)

    row_i = 4
    for sec in SECTION_ORDER:
        group = grouped.get(sec)
        if not group:
            continue
        head, alt, label = SECTION_COLORS[sec]
        ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=ncols)
        c = ws.cell(row=row_i, column=1, value=label)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.fill = _fill(head)
        c.alignment = Alignment(vertical="center", indent=1)
        row_i += 1
        for n, r in enumerate(group):
            v48 = "NO 48V" if r.get("ribbon") else ("✓" if r.get("phantom") else "")
            mic = r.get("mic") or ""
            if r.get("tour"):
                mic = f"⚑ {mic}"
            values = [
                r.get("ch"),
                r.get("name") or r.get("instrument") or "",
                mic,
                r.get("split") or r.get("port") or "",
                v48,
                r.get("stand") or "",
                r.get("notes") or "",
            ]
            for i, val in enumerate(values, start=1):
                cell = ws.cell(row=row_i, column=i, value=val)
                cell.border = BORDER
                cell.font = Font(name="Consolas" if i in (1, 4) else "Calibri", size=10)
                if i == 7:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif i in (1, 5):
                    cell.alignment = Alignment(horizontal="center")
                if r.get("tour"):
                    cell.fill = _fill("#FFF3CD")
                elif n % 2:
                    cell.fill = _fill(alt)
            if r.get("ribbon"):
                ws.cell(row=row_i, column=5).font = Font(name="Calibri", size=10, bold=True, color="B91C1C")
            elif r.get("phantom"):
                ws.cell(row=row_i, column=5).font = Font(name="Calibri", size=10, bold=True, color="065F46")
            row_i += 1

    _rig_info_tab(wb, sheet, console)

    # Outputs on a second tab
    ws2 = wb.create_sheet("Outputs & Power")
    for i, (label, width) in enumerate([("Bus", 14), ("Feeds", 24), ("Port", 16), ("Device", 22), ("Location", 16), ("Notes", 30)], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width
        c = ws2.cell(row=1, column=i, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = _fill("#111827")
    r_i = 2
    for o in sheet.get("outputs", []):
        if not (o.get("name") or o.get("port") or o.get("device")):
            continue
        for i, val in enumerate([o.get("bus"), o.get("name"), o.get("port"), o.get("device"), o.get("location"), o.get("notes")], start=1):
            cell = ws2.cell(row=r_i, column=i, value=val or "")
            cell.border = BORDER
            cell.font = Font(name="Consolas" if i in (1, 3) else "Calibri", size=10)
        r_i += 1

    r_i += 2
    for d in sheet.get("power", []):
        c = ws2.cell(row=r_i, column=1, value=f"{d.get('name','')}  {d.get('location','')}  {d.get('feed','')}".strip())
        c.font = Font(name="Calibri", size=11, bold=True)
        c.fill = _fill("#E5E7EB")
        r_i += 1
        for ck in d.get("circuits", []):
            for i, val in enumerate([ck.get("ckt"), ck.get("load"), ck.get("amps"), ck.get("notes")], start=1):
                cell = ws2.cell(row=r_i, column=i, value=val or "")
                cell.border = BORDER
                cell.font = Font(name="Consolas" if i == 1 else "Calibri", size=10)
            r_i += 1

    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
