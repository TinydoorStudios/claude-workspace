"""Read a filled-in template intake workbook back into the house templates.

    ../ShowBuilder/.venv/bin/python -m tools.import_intake "Patchbay Template Intake.xlsx" [--dry-run]

Each tab maps to the sheet whose id is in its subtitle line ("Sheet id: …"). Blocks
that contain at least one filled row replace that part of the sheet; a block left
entirely blank is left alone, so a partial pass is safe. Templates are unlocked for
the write and re-locked afterwards, and every sheet gets a revision snapshot first.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from backend.schema import (blank_channel, blank_console, blank_device,  # noqa: E402
                            blank_counts, migrate)
from backend.store import Store, new_id  # noqa: E402

TRUE = {"y", "yes", "true", "1", "on", "✓"}


def s(v) -> str:
    return "" if v is None else str(v).strip()


def num(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return default


def read_blocks(ws) -> tuple[str, dict[str, list[dict]]]:
    """Split a tab into {BLOCK: [row dicts]} plus the sheet id from the subtitle."""
    sheet_id = ""
    blocks: dict[str, list[dict]] = {}
    key = None
    header: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = [s(c) for c in row]
        first = cells[0] if cells else ""
        if first.startswith("Sheet id:"):
            sheet_id = first.split("Sheet id:")[1].split("·")[0].strip()
            continue
        if first.startswith("## "):
            key = first[3:].strip()
            blocks[key] = []
            header = []
            continue
        if key is None or not any(cells):
            continue
        if not header:
            header = cells
            continue
        if not any(cells):
            continue
        blocks[key].append({h: cells[i] if i < len(cells) else "" for i, h in enumerate(header) if h})
    return sheet_id, blocks


def filled(rows: list[dict]) -> list[dict]:
    return [r for r in rows if any(v for v in r.values())]


def apply_tab(sheet: dict, blocks: dict[str, list[dict]]) -> list[str]:
    notes = []

    loc_rows = filled(blocks.get("LOCATION", []))
    if loc_rows:
        by_label = {r.get("Field", ""): r.get("Value", "") for r in loc_rows}
        mapping = {
            "Project": ("location", "project"), "Client / org": ("location", "client"),
            "Site / venue name": ("location", "site"), "Room": ("location", "room"),
            "Address": ("location", "address"), "City": ("location", "city"),
            "State": ("location", "state"), "Zip": ("location", "zip"),
            "Venue label": (None, "venue_label"), "Default FOH": ("meta", "foh"),
            "Default MON": ("meta", "mon"), "Default show time": ("meta", "showtime"),
            "Rig notes": ("meta", "notes"),
        }
        for label, (parent, key) in mapping.items():
            if label not in by_label:
                continue
            value = by_label[label]
            if parent:
                sheet.setdefault(parent, {})[key] = value
            else:
                sheet[key] = value
        notes.append("location")

    con_rows = filled(blocks.get("CONSOLES", []))
    if con_rows:
        existing = {c.get("name", ""): c for c in sheet.get("consoles", [])}
        consoles = []
        for r in con_rows:
            name = r.get("Console name") or "Console"
            con = existing.get(name) or blank_console(r.get("Desk preset") or "q225", name)
            con["name"] = name
            con["preset"] = (r.get("Desk preset") or con.get("preset") or "q225").lower()
            con["manufacturer"] = r.get("Manufacturer", con.get("manufacturer", ""))
            con["model"] = r.get("Model", con.get("model", ""))
            con["fw"] = r.get("Firmware", con.get("fw", ""))
            con.setdefault("counts", blank_counts())
            for col, key in [("Inputs", "inputs"), ("Busses", "busses"), ("Auxes", "auxes"),
                             ("DCAs", "dcas"), ("Mute groups", "mutes"), ("Matrix", "matrix"),
                             ("Outputs", "outputs")]:
                if s(r.get(col)):
                    con["counts"][key] = num(r.get(col))
            con["network"] = {
                "ip": r.get("IP address", ""), "subnet": r.get("Subnet", ""),
                "gateway": r.get("Gateway", ""), "dns": r.get("DNS", ""),
            }
            con["tielines"] = r.get("Tie lines", "").lower() in TRUE
            # Channel rows follow the count, keeping whatever is already filled in.
            want = con["counts"].get("inputs") or len(con.get("channels", []))
            while len(con["channels"]) < want:
                con["channels"].append(blank_channel(len(con["channels"]) + 1))
            del con["channels"][want:]
            consoles.append(con)
        sheet["consoles"] = consoles
        sheet["console"] = consoles[0]["preset"]
        notes.append(f"{len(consoles)} console(s)")

    by_name = {c.get("name", ""): c for c in sheet["consoles"]}
    first = sheet["consoles"][0]

    conn_rows = filled(blocks.get("CONNECTIONS", []))
    if conn_rows:
        for con in sheet["consoles"]:
            con["connections"] = []
        for r in conn_rows:
            con = by_name.get(r.get("Console", ""), first)
            con["connections"].append({"id": new_id(), "name": r.get("Name", ""),
                                       "type": r.get("Type", ""), "channels": r.get("Channels", "")})
        notes.append(f"{len(conn_rows)} connection(s)")

    dev_rows = filled(blocks.get("DEVICES", []))
    if dev_rows:
        devices = []
        for r in dev_rows:
            dev = blank_device("network" if r.get("Kind (io/network)", "").lower().startswith("n") else "io")
            targets = [x.strip() for x in r.get("Assign to console", "").split(",") if x.strip()]
            dev.update({
                "name": r.get("Name", ""), "inputs": num(r.get("Inputs")), "outputs": num(r.get("Outputs")),
                "protocol": r.get("Protocol", ""), "ip": r.get("IP", ""), "location": r.get("Location", ""),
                "format": r.get("Format", ""), "notes": r.get("Notes", ""),
                "consoles": [by_name[t]["id"] for t in targets if t in by_name] or ([first["id"]] if dev["kind"] == "io" else []),
            })
            devices.append(dev)
        sheet["devices"] = devices
        notes.append(f"{len(devices)} device(s)")

    pos_rows = filled(blocks.get("STAGE POSITIONS", []))
    if pos_rows:
        positions: dict[str, dict] = {}
        for r in pos_rows:
            name = r.get("Position", "")
            pos = positions.setdefault(name, {"id": new_id(), "name": name,
                                              "note": r.get("Position note", ""), "runs": []})
            if not pos["note"]:
                pos["note"] = r.get("Position note", "")
            if any(s(r.get(k)) for k in ("Run label", "Device", "Port", "Run notes")):
                pos["runs"].append({"id": new_id(), "label": r.get("Run label", ""),
                                    "device": r.get("Device", ""), "port": r.get("Port", ""),
                                    "notes": r.get("Run notes", "")})
        sheet["positions"] = list(positions.values())
        notes.append(f"{len(positions)} stage position(s)")

    run_rows = filled(blocks.get("DATA RUNS", []))
    if run_rows:
        sheet["data_runs"] = [{"id": new_id(), "label": r.get("Label", ""), "type": r.get("Type", "")}
                              for r in run_rows]
        notes.append(f"{len(run_rows)} data run(s)")

    out_rows = filled(blocks.get("OUTPUTS", []))
    if out_rows:
        for con in sheet["consoles"]:
            con["outputs"] = []
        for r in out_rows:
            con = by_name.get(r.get("Console", ""), first)
            con["outputs"].append({"id": new_id(), "bus": r.get("Bus", ""), "name": r.get("Feeds", ""),
                                   "port": r.get("Port", ""), "device": r.get("Device", ""),
                                   "location": r.get("Location", ""), "notes": r.get("Notes", "")})
        notes.append(f"{len(out_rows)} output(s)")

    pwr_rows = filled(blocks.get("POWER", []))
    if pwr_rows:
        distros: dict[str, dict] = {}
        for r in pwr_rows:
            name = r.get("Distro", "")
            d = distros.setdefault(name, {"id": new_id(), "name": name, "location": r.get("Location", ""),
                                          "feed": r.get("Feed", ""), "circuits": []})
            if any(s(r.get(k)) for k in ("Circuit", "Load", "Amps", "Notes")):
                d["circuits"].append({"id": new_id(), "ckt": r.get("Circuit", ""), "load": r.get("Load", ""),
                                      "amps": r.get("Amps", ""), "notes": r.get("Notes", "")})
        sheet["power"] = list(distros.values())
        notes.append(f"{len(distros)} distro(s)")

    con_rows2 = filled(blocks.get("CONTACTS", []))
    if con_rows2:
        sheet["contacts"] = [{"id": new_id(), "name": r.get("Name", ""), "role": r.get("Role", ""),
                              "phone": r.get("Phone", ""), "email": r.get("Email", ""),
                              "notes": r.get("Notes", "")} for r in con_rows2]
        notes.append(f"{len(con_rows2)} contact(s)")

    ch_rows = filled(blocks.get("CHANNELS", []))
    if ch_rows:
        for con in sheet["consoles"]:
            for row in con["channels"]:
                row.update({"name": "", "mic": "", "port": "", "phantom": False, "stand": "",
                            "box": "", "split": "", "notes": ""})
        for i, r in enumerate(ch_rows):
            con = by_name.get(r.get("Console", ""), first)
            index = num(r.get("CH"), 0) - 1
            if index < 0 or index >= len(con["channels"]):
                con["channels"].append(blank_channel(num(r.get("CH"), len(con["channels"]) + 1)))
                index = len(con["channels"]) - 1
            row = con["channels"][index]
            row.update({
                "ch": num(r.get("CH"), row.get("ch") or index + 1),
                "name": r.get("Name", ""), "mic": r.get("Mic / DI", ""), "port": r.get("Port", ""),
                "phantom": r.get("48V (y/n)", "").lower() in TRUE, "stand": r.get("Stand", ""),
                "section": r.get("Section") or row.get("section") or "SPARE",
                "box": r.get("Device", ""), "split": r.get("Split", ""), "notes": r.get("Notes", ""),
            })
        for con in sheet["consoles"]:
            con["counts"]["inputs"] = len(con["channels"])
        notes.append(f"{len(ch_rows)} channel(s)")

    return notes


def main(path: str, dry_run: bool = False) -> None:
    wb = load_workbook(path, data_only=True)
    store = Store()
    for name in wb.sheetnames:
        if name == "Read Me":
            continue
        sheet_id, blocks = read_blocks(wb[name])
        if not sheet_id:
            print(f"! {name}: no sheet id in the subtitle — skipped")
            continue
        if not store.exists(sheet_id):
            print(f"! {name}: no sheet '{sheet_id}' on disk — skipped")
            continue
        sheet = migrate(store.get(sheet_id))
        was_locked = bool(sheet.get("locked"))
        sheet["locked"] = False
        changed = apply_tab(sheet, blocks)
        sheet["locked"] = was_locked
        if dry_run:
            print(f"[dry run] {sheet_id}: {', '.join(changed) or 'nothing filled in'}")
            continue
        store.save(sheet_id, sheet, bump=True)
        print(f"{sheet_id}: {', '.join(changed) or 'nothing filled in'}"
              + (" (re-locked)" if was_locked else ""))


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--dry-run"]
    if not args:
        print(__doc__)
        raise SystemExit(1)
    main(args[0], dry_run="--dry-run" in sys.argv)
