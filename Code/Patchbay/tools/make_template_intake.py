"""Build the template intake workbook — one tab per house rig, every field Patchbay
stores, pre-filled with what the templates already know.

    ../ShowBuilder/.venv/bin/python -m tools.make_template_intake

Writes "Patchbay Template Intake.xlsx" next to the repo. Fill in the blanks (and
correct anything marked CONFIRM), hand it back, then:

    ../ShowBuilder/.venv/bin/python -m tools.import_intake "Patchbay Template Intake.xlsx"

Block layout is what the importer parses: a "## BLOCK" marker in column A, then a
header row, then rows until the next marker. Keep that shape and add rows freely.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from backend.store import Store  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "Patchbay Template Intake.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="1F2937")
BLOCK_FILL = PatternFill("solid", fgColor="374151")
HEAD_FILL = PatternFill("solid", fgColor="111827")
FILL_ME = PatternFill("solid", fgColor="FFF3CD")
THIN = Border(bottom=Side(style="thin", color="D1D5DB"))

LOCATION_FIELDS = [
    ("Project", "location.project", "Prints as the rig-information Project line"),
    ("Client / org", "location.client", "e.g. 3CDC, Jazz At The Memo"),
    ("Site / venue name", "location.site", ""),
    ("Room", "location.room", "FOH, amp room, plaza…"),
    ("Address", "location.address", ""),
    ("City", "location.city", ""),
    ("State", "location.state", ""),
    ("Zip", "location.zip", ""),
    ("Venue label", "venue_label", "Short name used in the sub-bar"),
    ("Default FOH", "meta.foh", "Leave blank if it changes per show"),
    ("Default MON", "meta.mon", ""),
    ("Default show time", "meta.showtime", ""),
    ("Rig notes", "meta.notes", "Room behaviour, standing rules, gotchas"),
]

CONSOLE_COLS = ["Console name", "Desk preset", "Manufacturer", "Model", "Firmware",
                "Inputs", "Busses", "Auxes", "DCAs", "Mute groups", "Matrix", "Outputs",
                "IP address", "Subnet", "Gateway", "DNS", "Tie lines"]
CONNECTION_COLS = ["Console", "Name", "Type", "Channels"]
DEVICE_COLS = ["Kind (io/network)", "Name", "Inputs", "Outputs", "Protocol", "IP",
               "Location", "Format", "Notes", "Assign to console"]
POSITION_COLS = ["Position", "Position note", "Run label", "Device", "Port", "Run notes"]
DATARUN_COLS = ["Label", "Type"]
OUTPUT_COLS = ["Console", "Bus", "Feeds", "Port", "Device", "Location", "Notes"]
POWER_COLS = ["Distro", "Location", "Feed", "Circuit", "Load", "Amps", "Notes"]
CONTACT_COLS = ["Name", "Role", "Phone", "Email", "Notes"]
CHANNEL_COLS = ["Console", "CH", "Name", "Mic / DI", "Port", "48V (y/n)", "Stand",
                "Section", "Device", "Split", "Notes"]

SPARE = {"consoles": 1, "connections": 4, "devices": 4, "positions": 6, "data_runs": 4,
         "outputs": 6, "power": 6, "contacts": 5, "channels": 8}


class Tab:
    def __init__(self, ws, title: str, subtitle: str):
        self.ws = ws
        self.row = 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        c.fill = TITLE_FILL
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 30
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        c = ws.cell(row=2, column=1, value=subtitle)
        c.font = Font(name="Calibri", size=10, color="FFFFFF")
        c.fill = BLOCK_FILL
        c.alignment = Alignment(vertical="center", indent=1)
        self.row = 4

    def block(self, key: str, help_text: str, columns: list[str], rows: list[list], spare: int) -> None:
        ws = self.ws
        ws.cell(row=self.row, column=1, value=f"## {key}").font = Font(name="Consolas", size=11, bold=True)
        if help_text:
            c = ws.cell(row=self.row, column=2, value=help_text)
            c.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
        self.row += 1
        for i, label in enumerate(columns, start=1):
            c = ws.cell(row=self.row, column=i, value=label)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = HEAD_FILL
            c.alignment = Alignment(horizontal="center")
        self.row += 1
        for values in rows:
            for i, v in enumerate(values, start=1):
                c = ws.cell(row=self.row, column=i, value=v if v not in (None, "") else "")
                c.font = Font(name="Calibri", size=10)
                c.border = THIN
                if v in (None, ""):
                    c.fill = FILL_ME
            self.row += 1
        for _ in range(spare):
            for i in range(1, len(columns) + 1):
                c = ws.cell(row=self.row, column=i, value="")
                c.border = THIN
                c.fill = FILL_ME
            self.row += 1
        self.row += 1

    def widths(self, widths: dict[int, int]) -> None:
        for col, w in widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = w


def readme(wb: Workbook) -> None:
    ws = wb.create_sheet("Read Me", 0)
    ws.column_dimensions["A"].width = 110
    lines = [
        ("Patchbay — template intake", 18, True),
        ("", 10, False),
        ("One tab per house rig. Fill in the blanks, correct anything that's wrong, then hand the file back.", 11, False),
        ("Amber cells are the ones waiting on you. Anything already filled in is what the template holds today.", 11, False),
        ("", 10, False),
        ("Rules of the sheet:", 12, True),
        ("• Don't rename the tabs, the '## BLOCK' markers in column A, or the header rows — the importer reads those.", 11, False),
        ("• Add as many rows as you need inside a block. Blank rows are ignored.", 11, False),
        ("• Delete a row you don't want. Leaving it blank does the same thing.", 11, False),
        ("• 48V column takes y / n. Tie lines takes y / n. Kind takes io or network.", 11, False),
        ("• Assign to console: the console name from the CONSOLES block. Blank = the first console.", 11, False),
        ("• Desk preset is one of q225, m32, wing — that's what drives the port surface in Easy Patch.", 11, False),
        ("", 10, False),
        ("What each block feeds:", 12, True),
        ("LOCATION — the rig-information block at the top of the printed patch sheet.", 11, False),
        ("CONSOLES — Console Info: make, model, counts, networking. One row per desk on that rig.", 11, False),
        ("CONNECTIONS — the I/O slots on a console (Dante, MADI, AES, local XLR).", 11, False),
        ("DEVICES — stage racks, splits and network gear. I/O devices become patch columns in Easy Patch.", 11, False),
        ("STAGE POSITIONS — stage positions and their runs. Repeat the position name on each of its runs.", 11, False),
        ("DATA RUNS — Cat6 / fibre drops on stage.", 11, False),
        ("OUTPUTS — busses, what they feed, the output port and the box on the other end.", 11, False),
        ("POWER — distros and circuits. Repeat the distro name on each of its circuits.", 11, False),
        ("CONTACTS — people who print on the sheet.", 11, False),
        ("CHANNELS — permanent channels only: the Memo crowd rig, house DIs, anything patched year-round.", 11, False),
        ("           Leave the CH number blank if the show decides it.", 11, False),
        ("", 10, False),
        ("Importing replaces the blocks you filled in and re-locks the template. Sheets already cloned off a", 11, False),
        ("template don't change — they're their own copies.", 11, False),
    ]
    for i, (text, size, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold)


def build() -> Path:
    store = Store()
    wb = Workbook()
    wb.remove(wb.active)

    for row in store.list():
        # House rigs only — the locked templates. Show sheets aren't intake material.
        if row["kind"] != "install" or not row.get("locked"):
            continue
        sheet = store.get(row["id"])
        ws = wb.create_sheet(sheet["name"][:31].replace("/", "-"))
        tab = Tab(ws, sheet["name"], f"Sheet id: {sheet['id']}  ·  fill the amber cells  ·  don't rename the ## blocks")
        tab.widths({1: 24, 2: 26, 3: 20, 4: 18, 5: 16, 6: 14, 7: 14, 8: 16, 9: 22, 10: 22, 11: 26})

        loc = sheet.get("location", {})
        meta = sheet.get("meta", {})
        def val(path):
            head, _, tail = path.partition(".")
            return (sheet.get(head, {}) or {}).get(tail, "") if tail else sheet.get(head, "")
        tab.block("LOCATION", "one row per field — put the answer in Value",
                  ["Field", "Value", "Notes"],
                  [[label, val(path), note] for label, path, note in LOCATION_FIELDS], 0)

        consoles = sheet.get("consoles", [])
        tab.block("CONSOLES", "one row per desk on this rig", CONSOLE_COLS,
                  [[c.get("name", ""), c.get("preset", ""), c.get("manufacturer", ""), c.get("model", ""),
                    c.get("fw", ""), c["counts"].get("inputs", ""), c["counts"].get("busses", ""),
                    c["counts"].get("auxes", ""), c["counts"].get("dcas", ""), c["counts"].get("mutes", ""),
                    c["counts"].get("matrix", ""), c["counts"].get("outputs", ""),
                    c["network"].get("ip", ""), c["network"].get("subnet", ""), c["network"].get("gateway", ""),
                    c["network"].get("dns", ""), "y" if c.get("tielines") else "n"] for c in consoles],
                  SPARE["consoles"])

        tab.block("CONNECTIONS", "console I/O slots — Dante, MADI, AES, local XLR", CONNECTION_COLS,
                  [[c.get("name", ""), k.get("name", ""), k.get("type", ""), k.get("channels", "")]
                   for c in consoles for k in c.get("connections", [])], SPARE["connections"])

        tab.block("DEVICES", "stage racks, splits, network gear", DEVICE_COLS,
                  [[d.get("kind", "io"), d.get("name", ""), d.get("inputs", ""), d.get("outputs", ""),
                    d.get("protocol", ""), d.get("ip", ""), d.get("location", ""), d.get("format", ""),
                    d.get("notes", ""),
                    ", ".join(c.get("name", "") for c in consoles if c["id"] in (d.get("consoles") or []))]
                   for d in sheet.get("devices", [])], SPARE["devices"])

        tab.block("STAGE POSITIONS", "repeat the position name on each of its runs", POSITION_COLS,
                  [[p.get("name", ""), p.get("note", ""), r.get("label", ""), r.get("device", ""),
                    r.get("port", ""), r.get("notes", "")]
                   for p in sheet.get("positions", []) for r in (p.get("runs") or [{}])], SPARE["positions"])

        tab.block("DATA RUNS", "Cat6 / fibre drops on stage", DATARUN_COLS,
                  [[d.get("label", ""), d.get("type", "")] for d in sheet.get("data_runs", [])], SPARE["data_runs"])

        tab.block("OUTPUTS", "busses and what they feed", OUTPUT_COLS,
                  [[c.get("name", ""), o.get("bus", ""), o.get("name", ""), o.get("port", ""),
                    o.get("device", ""), o.get("location", ""), o.get("notes", "")]
                   for c in consoles for o in c.get("outputs", [])], SPARE["outputs"])

        tab.block("POWER", "repeat the distro name on each of its circuits", POWER_COLS,
                  [[d.get("name", ""), d.get("location", ""), d.get("feed", ""), k.get("ckt", ""),
                    k.get("load", ""), k.get("amps", ""), k.get("notes", "")]
                   for d in sheet.get("power", []) for k in (d.get("circuits") or [{}])], SPARE["power"])

        tab.block("CONTACTS", "people who print on the sheet", CONTACT_COLS,
                  [[c.get("name", ""), c.get("role", ""), c.get("phone", ""), c.get("email", ""), c.get("notes", "")]
                   for c in sheet.get("contacts", [])], SPARE["contacts"])

        tab.block("CHANNELS", "permanent channels only — leave CH blank if the show decides it", CHANNEL_COLS,
                  [[c.get("name", ""), ch.get("ch", ""), ch.get("name", ""), ch.get("mic", ""), ch.get("port", ""),
                    "y" if ch.get("phantom") else "n", ch.get("stand", ""), ch.get("section", ""),
                    ch.get("box", ""), ch.get("split", ""), ch.get("notes", "")]
                   for c in consoles for ch in c.get("channels", []) if ch.get("name") or ch.get("mic")],
                  SPARE["channels"])

    readme(wb)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(build())
