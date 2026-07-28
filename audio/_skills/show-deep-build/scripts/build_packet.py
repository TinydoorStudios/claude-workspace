#!/usr/bin/env python3
"""
build_packet.py — deep-research show packet engine (show-deep-build skill).

ONE source of truth (a deep-research spec.json) -> every paperwork output:
  <Show> - FOH Channel Processing.md   (patcher input; feed to the venue .ses patcher)
  <Show> - Input List.xlsx             (+ Monitors / Reverbs sheets when the spec carries them)
  <Show> - Show Packet.pdf             (via show-packet-builder-template.py)
  <Show> - FOH EQ Reasoning.pdf        (the EQ Rationale — required deep-research deliverable)
  <Show> - MASTER.pdf                  (everything in one PDF: a clickable QUICK LINKS page,
                                        then packet + rationale + any band-provided
                                        "<Show> - Stage Plot.pdf"/"- Rider.pdf" found in the
                                        show folder; individual files still ship)

MASTER navigation (2026-07-27): page 1 of the MASTER is a quick-links index — every document,
every EQ section, and every channel is a clickable jump, and the same map is written as PDF
bookmarks for the reader's sidebar. Page numbers come from real page marks recorded while the
PDFs render, not from guessing.

House rules enforced at validation (2026-07-08 additions):
  - RESERVED template channels error out (fsq ch 10 = SNARE PL8 return; OH pair is
    STEREO on fader 9 — never split across 9/10)
  - WIRELESS faders (2026-07-26): fsq 33-36 / memo 41-44 = Wireless 1-4. A mic naming
    a wireless with no unit number ("W58") errors — never auto-assign a pack. A named
    wireless mults: the named input keeps its channel AND the wireless fader stays
    listed (warning if the fader row is missing, or if a non-wireless source is
    parked on one).
  - "reverbs" is REQUIRED every show, FSQ included (3 complementary vocal + 1-2
    instrument + general when warranted; Seventh Heaven presets verbatim, each with
    settings + in-plugin EQ + why, plus spec-level "reverb_pairing"). Explicit
    opt-out: "no_reverb": true.
  - Stage plots are BAND-PROVIDED — never generated; drop theirs in the show folder
    as "<Show> - Stage Plot.pdf" and the MASTER picks it up.

The .ses itself is built per-venue by the existing patcher (Q225 SES Patcher SOP) using the .md
this script writes. This script does NOT touch the .ses.

The spec is VALIDATED before anything is written (house rules: ribbon = no 48V, vocals cuts-only,
whole dB, band/freq ranges) and the .md is auto-linted with audio/_shared/md_lint.py after writing.
Validation errors abort; warnings print.

Usage:
  python3 build_packet.py --spec "<spec.json>" --out "<show folder>" \
      [--packet-builder "~/Documents/Claude/audio/show-packet-builder-template.py"]

Deps: openpyxl + reportlab (pip3 install --user; both present on the Mac as of 2026-07-06).
Spec schema: see references/spec-schema.md. Channels list only ACTIVE bands; any band 1-4 not
present is treated as FLAT. Band numbering is Brian's console convention: b1=low .. b4=high.
"""
import argparse, json, os, re, sys, importlib.util

VENUE_LABELS = {"fsq": "Fountain Square", "memo": "Memorial Hall", "wp": "Washington Park",
                "esp": "Elm Street Plaza", "csp": "Court Street Plaza", "zp": "Zeigler Park",
                "ia": "Imagination Alley", "greaves": "Greaves Concert Hall"}
KNOWN_SECTIONS = {"DRUMS", "BASS", "RHYTHM", "GUITAR", "KEYS", "PIANO", "STRINGS", "HORNS",
                  "VOCALS", "AMBIENT", "SPARE"}
RIBBON_FLAG = "⚠ RIBBON — NO 48V"
TOUR_FLAG = "⚑ TOUR — confirm at load-in"

# Template faders that are NOT inputs — assigning a source here clobbers an
# FX return (learned the hard way: Hot Magnolias put OH R on FSQ ch 10 and
# overwrote the snare plate return, 2026-07-08). The venue patcher also hard
# refuses these; catching it at spec time is cheaper.
RESERVED_CH = {
    "fsq": {10: "SNARE PL8 — snare plate reverb return. Overheads are STEREO "
                "on fader 9 (both OH mics on that one fader); an OH pair "
                "never spills onto 10."},
}

# House wireless receivers live on fixed template faders (Brian, 2026-07-26).
# Fill in a wireless 1-4 row and it lands on its own fader; when another input's
# mic names a wireless unit, the receiver is MULTED — the named input keeps its
# channel AND the wireless fader stays listed. Confirmed against both patcher
# templates' surface labels (FSQ 33-36, Memo 41-44).
WIRELESS_CH = {
    "fsq":  {1: 33, 2: 34, 3: 35, 4: 36},
    "memo": {1: 41, 2: 42, 3: 43, 4: 44},
}
# "Wireless 2" / "WL2" / "W2" / "W58 2" -> unit 2. Bare "W58"/"wireless" -> ask.
_WL_UNIT_RE = [re.compile(p, re.I) for p in (
    r"wireless\s*#?\s*([1-4])\b",
    r"\bwl\s*#?\s*([1-4])\b",
    r"\bw\s*58\s*[-/ ]?\s*([1-4])\b",
    r"\bw\s*#?\s*([1-4])\b",
)]
_WL_BARE_RE = re.compile(r"\b(w\s*58|wireless|wl)\b", re.I)

def wireless_unit(text):
    """(unit|None, is_wireless). unit None + True = named a wireless, no number."""
    s = str(text or "")
    for rx in _WL_UNIT_RE:
        m = rx.search(s)
        if m:
            return int(m.group(1)), True
    return None, bool(_WL_BARE_RE.search(s))

# ---------------------------------------------------------------- spec helpers
def load_spec(p):
    with open(p) as f:
        return json.load(f)

def venue_label(spec):
    return spec.get("venue_label") or VENUE_LABELS.get(str(spec.get("venue", "")).lower(), "")

def fmt_freq(f):
    f = float(f)
    return f"{f/1000:g} kHz" if f >= 1000 else f"{f:g} Hz"

def bands_by_num(ch):
    """Return {1..4: band-dict or None}. Missing / gain==None / type FLAT -> None."""
    out = {1: None, 2: None, 3: None, 4: None}
    for b in ch.get("bands", []):
        n = int(b["b"])
        if b.get("gain") is None or str(b.get("type", "")).upper() == "FLAT":
            continue
        out[n] = b
    return out

def patch_label(ch):
    return ch.get("patch") or f"Local {ch['ch']}"

def flag_notes(ch):
    """Channel notes with the ribbon/TOUR flags prepended (input list + packet rows)."""
    flags = []
    if ch.get("ribbon"):
        flags.append(RIBBON_FLAG)
    if ch.get("tour"):
        flags.append(TOUR_FLAG)
    base = ch.get("notes") or ""
    return " · ".join(flags + ([base] if base else [])) or None

def esc(s):
    """Escape user text for reportlab Paragraph markup."""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

# ------------------------------------------------- prose / research formatting
# The research block used to render as one wall of text. It now breaks into
# structured pieces (Brian, 2026-07-27): enumerated prose gets real paragraph
# breaks, TRACE lines get one line per layer, and a structured "research"
# object renders as a per-unit table instead of a paragraph.

_ENUM_BREAK = re.compile(r"(?<=[.;])\s+(?=(?:\(\d\)|First,|Second,|Third,|Fourth,|Fifth,|Finally,))")
_TRACE_SPLIT = re.compile(r"\bTRACE:\s*", re.I)
_LAYER_RE = re.compile(r"^\s*([A-Za-z][\w /]*?)\s*\((.*)\)\s*$", re.S)
TRACE_LAYERS = ("base", "equip", "genre", "artist", "venue")


def prose_html(text):
    """Escaped text with paragraph breaks at enumerated turns — no more wall."""
    parts = [p.strip() for p in _ENUM_BREAK.split(str(text or "")) if p.strip()]
    return "<br/><br/>".join(esc(p) for p in parts)


def trace_html(text, indent=""):
    """A TRACE body ('base(...) · equip(...) · …') as one bolded line per layer."""
    segs, depth, buf = [], 0, ""
    for chpart in str(text or ""):
        if chpart == "(":
            depth += 1
        elif chpart == ")":
            depth = max(0, depth - 1)
        if chpart in "·|" and depth == 0:
            segs.append(buf); buf = ""
        else:
            buf += chpart
    segs.append(buf)
    lines = []
    for seg in (s.strip(" .") for s in segs):
        if not seg:
            continue
        m = _LAYER_RE.match(seg)
        if m:
            lines.append(f"{indent}<b>{esc(m.group(1).lower())}</b> &nbsp;{esc(m.group(2))}")
        else:
            lines.append(f"{indent}{esc(seg)}")
    return "<br/>".join(lines)


def research_chunks(summary):
    """Legacy free-text research_summary -> readable per-unit rows.

    Splits on the lead-ins the deep build actually writes (GENRE VERIFIED, THE GIG,
    WEATHER, PER-UNIT, CH<n>, RECONCILIATION) so a 15 kB paragraph becomes a
    scannable table, pulls the mic out of the unit heading, chips the
    AGREE/DISAGREE/THIN verdict, and puts each TRACE layer on its own line.
    Returns [(head, mic, verdict, body_html), ...].
    """
    txt = re.sub(r"\s+", " ", str(summary or "")).strip()
    if not txt:
        return []
    pat = re.compile(r"(?=(?:CH\s?\d|GENRE VERIFIED|THE GIG|WEATHER|PER-UNIT|"
                     r"RECONCILIATION|SOURCES\b|KB WRITE))")
    raw = [c.strip() for c in pat.split(txt) if c.strip()]
    out = []
    for chunk in raw:
        head = mic = ""
        m = re.match(r"^(CH\s?[\d/–\-]+(?:\s*[/&]\s*[\d–\-]+)*[^—]{0,40}?)"
                     r"\s+x\s+(.{3,60}?)\s*—\s*", chunk)
        if m:
            head, mic, chunk = m.group(1).strip(), m.group(2).strip(), chunk[m.end():]
        else:
            m = re.match(r"^(CH\s?[\d/–\-]+(?:\s*[/&]\s*[\d–\-]+)*[^—:.]{0,40}?)\s*(?:—|:)\s*", chunk)
            if m:
                head, chunk = m.group(1).strip(), chunk[m.end():]
            else:
                m2 = re.match(r"^([A-Z][A-Z \-/]{2,40}?)\s*(?::|—|\()", chunk)
                if m2:
                    head = m2.group(1).strip()
                    chunk = chunk[m2.end(1):].lstrip(" —:")
        if _TRACE_SPLIT.search(chunk):
            body, trace = _TRACE_SPLIT.split(chunk, maxsplit=1)
        else:
            body, trace = chunk, ""
        vm = re.search(r"\b(AGREE|DISAGREE|THIN)\b", body)
        verdict = vm.group(1) if vm else ""
        body_html = esc(body.strip())
        if trace:
            body_html += ("<br/><font color='#0F3460'><b>TRACE</b></font><br/>"
                          + trace_html(trace))
        out.append((head, mic, verdict, body_html))
    return out

# ---------------------------------------------------------------- 0. validation
def validate_spec(spec):
    """House rules enforced as code. Returns (errors, warnings)."""
    errors, warnings = [], []
    if not spec.get("channels"):
        errors.append("spec has no channels")
        return errors, warnings
    reserved = RESERVED_CH.get(str(spec.get("venue", "")).lower(), {})
    # Reverb suggestions are a required deliverable, every show, FSQ included
    # (Brian, 2026-07-08): 3 complementary vocal options, 1-2 instrument, and
    # a general verb when warranted — Seventh Heaven Pro, presets verbatim
    # from the reverb KB, each with settings + in-plugin EQ + why, plus a
    # spec-level "reverb_pairing" line on using them together.
    revs = spec.get("reverbs") or []
    if not revs and not spec.get("no_reverb"):
        errors.append('spec has no "reverbs" — reverb suggestions are required every '
                      'show (FSQ included). 3 vocal + 1-2 instrument (+ general), '
                      'Seventh Heaven presets verbatim, settings + plugin EQ + why. '
                      'If Brian explicitly wants none, set "no_reverb": true.')
    if revs:
        roles = [str(r.get("role", "")).lower() for r in revs if isinstance(r, dict)]
        if roles.count("vocal") < 3:
            warnings.append(f"reverbs: {roles.count('vocal')} vocal option(s) — Brian wants "
                            "3 complementary vocal choices")
        if roles.count("instrument") < 1:
            warnings.append("reverbs: no instrument option — Brian wants 1-2 (horn-specific when asked)")
        if not spec.get("reverb_pairing"):
            warnings.append('reverbs present but no "reverb_pairing" — add the how-they-work-'
                            'together line')
        for r in revs:
            if isinstance(r, dict) and not (r.get("plugin_eq") or r.get("settings")):
                warnings.append(f"reverb {r.get('preset','?')!r}: no settings/plugin_eq — include "
                                "the in-plugin moves, not just the preset name")
    # Research is a required deliverable and it has to be READABLE (2026-07-27):
    # the structured "research" object renders as the per-unit table. A legacy
    # free-text research_summary still builds — it gets chunked — but warns.
    R = spec.get("research")
    if isinstance(R, dict):
        if not R.get("genre_verified"):
            warnings.append('research: no "genre_verified" — the genre is verified with named '
                            "evidence before any research runs")
        units = R.get("units") or []
        if not units:
            warnings.append('research: no "units" — every researched instrument x mic gets a row '
                            "(capsule fact + external source + verdict + TRACE)")
        for u in units:
            tag = f"research unit {u.get('ch', '?')} ({u.get('source', '?')})"
            if not u.get("finding"):
                warnings.append(f"{tag}: no finding — needs the quantitative capsule fact")
            if not u.get("sources"):
                warnings.append(f"{tag}: no external source named — the KB alone is not research")
            v = str(u.get("verdict", "")).upper()
            if v not in ("AGREE", "DISAGREE", "THIN"):
                warnings.append(f"{tag}: verdict {u.get('verdict')!r} — one word, "
                                "AGREE / DISAGREE / THIN")
            tr = u.get("trace")
            if isinstance(tr, dict):
                missing = [k for k in TRACE_LAYERS if not tr.get(k)]
                if missing:
                    warnings.append(f"{tag}: TRACE missing {', '.join(missing)} — every layer "
                                    'carries a value or an explicit "no change"')
            elif not tr:
                warnings.append(f"{tag}: no TRACE")
        if not R.get("reconciliation"):
            warnings.append('research: no "reconciliation" — state the web-vs-KB forks or '
                            '"no web/KB disagreements"')
    elif spec.get("research_summary"):
        warnings.append('research: free-text "research_summary" only — new builds should write '
                        'the structured "research" object (genre_verified / gig / conditions / '
                        "units[] / reconciliation); it renders as the readable per-unit table "
                        "instead of chunked prose")
    else:
        warnings.append("spec has no research block — the deep build's research is a required "
                        "deliverable")
    wl_map = WIRELESS_CH.get(str(spec.get("venue", "")).lower(), {})
    wl_faders = {fader: unit for unit, fader in wl_map.items()}
    wl_named = {}          # unit -> [tags that call for it]
    seen_ch, seen_secs, last_sec = set(), set(), None
    for ch in spec["channels"]:
        cid = ch.get("ch")
        tag = f"Ch {cid} ({ch.get('name','?')})"
        # --- wireless (Brian, 2026-07-26) -------------------------------
        unit, is_wl = wireless_unit(f"{ch.get('mic','')} {ch.get('name','')}")
        if wl_map:
            if unit:
                wl_named.setdefault(unit, []).append(tag)
            elif is_wl:
                errors.append(f"{tag}: mic names a wireless with no unit number "
                              f"({ch.get('mic')!r}) — never auto-assign a unit; "
                              "ask Brian which wireless (1-4) it is")
            if cid in wl_faders and unit and unit != wl_faders[cid]:
                errors.append(f"{tag}: Ch {cid} is the Wireless {wl_faders[cid]} fader on the "
                              f"{spec.get('venue','')} template but the mic names Wireless {unit}")
            elif cid in wl_faders and not is_wl:
                warnings.append(f"{tag}: Ch {cid} is the Wireless {wl_faders[cid]} fader on the "
                                f"{spec.get('venue','')} template — parking a non-wireless source "
                                "here buries the receiver's home channel")
        if cid in seen_ch:
            errors.append(f"{tag}: duplicate channel number")
        seen_ch.add(cid)
        if cid in reserved:
            errors.append(f"{tag}: Ch {cid} is RESERVED on the "
                          f"{spec.get('venue','')} template — {reserved[cid]}")
        if not ch.get("mic"):
            errors.append(f"{tag}: no mic/DI — don't guess a mic from an instrument")
        if not ch.get("instrument"):
            warnings.append(f"{tag}: no instrument field (input list falls back to fader name)")
        if ch.get("ribbon") and ch.get("phantom"):
            errors.append(f"{tag}: ribbon with phantom=true — NO 48V on ribbons, ever")
        sec = str(ch.get("section", "")).upper()
        if sec not in KNOWN_SECTIONS:
            warnings.append(f"{tag}: unknown section {sec!r}")
        if sec != last_sec:
            if sec in seen_secs:
                warnings.append(f"{tag}: section {sec} reappears out of order — "
                                "packet/rationale section bars will repeat; group the channels")
            seen_secs.add(sec)
            last_sec = sec
        hpf = ch.get("hpf", 20)
        if not (20 <= float(hpf) <= 20000):
            errors.append(f"{tag}: HPF {hpf} outside 20..20000")
        lpf = ch.get("lpf")
        if lpf is not None and not (20 <= float(lpf) <= 20000):
            errors.append(f"{tag}: LPF {lpf} outside 20..20000")
        for b in ch.get("bands", []):
            n = b.get("b")
            if n not in (1, 2, 3, 4):
                errors.append(f"{tag}: band number {n!r} not 1..4")
                continue
            g = b.get("gain")
            if g is None or str(b.get("type", "")).upper() == "FLAT":
                continue
            if not (-18 <= g <= 18):
                errors.append(f"{tag} B{n}: gain {g} outside ±18")
            if g != int(g):
                warnings.append(f"{tag} B{n}: fractional gain {g} dB (house rule: whole dB)")
            f = b.get("freq")
            if f is None or not (20 <= float(f) <= 20000):
                errors.append(f"{tag} B{n}: freq {f!r} outside 20..20000")
            q = b.get("q")
            if q is None or not (0.3 <= float(q) <= 20):
                warnings.append(f"{tag} B{n}: Q {q!r} outside 0.3..20")
            if sec == "VOCALS" and g > 0 and not b.get("approved"):
                errors.append(f"{tag} B{n}: +{g} dB boost on a VOCAL — vocals are cuts-only, "
                              "every genre (feedback control). If Brian explicitly approved it, "
                              'add "approved": true to the band.')
            if n == 4 and str(b.get("type", "")).upper() == "SHELF" and g > 0 and not b.get("approved"):
                warnings.append(f"{tag} B4: high-shelf boost — house rule is no high shelf "
                                "unless Brian asked; confirm it was requested")
            d = b.get("deq")
            if d and not all(k in d for k in ("thr", "atk_ms", "rel_ms")):
                errors.append(f"{tag} B{n}: DEQ missing thr/atk_ms/rel_ms")
        name = str(ch.get("name", ""))
        if len(name) > 12:
            warnings.append(f"{tag}: fader name is {len(name)} chars (>12 — legibility)")
    # Every wireless a channel calls for keeps its own fader listed too — the
    # receiver is multed to both (Brian, 2026-07-26).
    for unit, tags in sorted(wl_named.items()):
        fader = wl_map[unit]
        if fader not in seen_ch:
            warnings.append(f"Wireless {unit} is called for by {', '.join(tags)} but Ch {fader} "
                            f"(its {spec.get('venue','')} template fader) isn't in the spec — the "
                            "receiver mults to both; list the wireless channel as well")
    return errors, warnings

# ---------------------------------------------------------------- 1. .md (patcher input)
def write_md(spec, folder):
    show = spec["show_name"]
    vl = venue_label(spec)
    sub = " · ".join(x for x in [vl, spec.get("console_label", "DiGiCo Quantum 225"),
                                 spec.get("show_date", ""), spec.get("rev", "Rev 1.0")] if x)
    L = [f"# {show} — FOH Channel Processing", f"## {sub}",
         "*Deep-research pass. Active channels only. Band order: B4 (high) -> B3 -> B2 -> B1 (low).*", ""]
    for ch in spec["channels"]:
        L.append(f"## Ch {ch['ch']} | {ch['name']} | {ch['mic']}")
        lpf = ch.get("lpf")
        L.append(f"HPF: {ch.get('hpf',20):g} | LPF: {'OFF' if not lpf else f'{lpf:g}'}")
        bb = bands_by_num(ch)
        for n in (4, 3, 2, 1):
            b = bb[n]
            if not b:
                L.append(f"B{n}: FLAT"); continue
            ty = "SHELF" if str(b["type"]).upper() == "SHELF" else "BELL"
            line = f"B{n}: {b['gain']:+g} | {b['freq']:g} | {b['q']:g} | {ty}"
            d = b.get("deq")
            if d:
                line += f" | DEQ: thr={d['thr']:g} atk={d['atk_ms']:g}ms rel={d['rel_ms']:g}ms"
            L.append(line)
        L.append("")
    p = os.path.join(folder, f"{show} - FOH Channel Processing.md")
    open(p, "w").write("\n".join(L)); return p

def lint_md(md_path):
    """Auto-lint the written .md with the shared linter. Errors abort the build."""
    shared = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                           "..", "..", "..", "_shared"))
    if not os.path.isdir(shared):
        print("note: _shared not found — md_lint skipped"); return
    sys.path.insert(0, shared)
    try:
        import md_lint
    except ImportError:
        print("note: md_lint not importable — skipped"); return
    errors, warnings = md_lint.lint(md_path)
    for w in warnings:
        print("  md-lint warn:", w)
    if errors:
        for e in errors:
            print("  md-lint ERROR:", e)
        raise SystemExit("md_lint failed — the written .md is not patcher-safe")
    print("md_lint PASS —", len(warnings), "warnings")

# ---------------------------------------------------------------- 2. Input List .xlsx
def write_xlsx(spec, folder):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    show = spec["show_name"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Input List"
    ws.append([show] + [None]*6)
    ws.append([f"{venue_label(spec) or spec.get('console_label','')} · {spec.get('show_date','')} · {spec.get('rev','')}"] + [None]*6)
    ws.append(["Ch", "Instrument", "Mic / DI", "Patch", "48V", "Stand", "Notes"])
    for c in ws[3]:
        c.font = Font(bold=True)
    red = Font(color="9B2222", bold=True)
    for ch in spec["channels"]:
        ws.append([ch["ch"], ch.get("instrument") or ch["name"], ch["mic"], patch_label(ch),
                   "✓" if ch.get("phantom") else None, ch.get("stand", "—"), flag_notes(ch)])
        if ch.get("ribbon"):
            ws[ws.max_row][6].font = red
    ws[1][0].font = Font(bold=True, size=14); ws[2][0].font = Font(italic=True, size=10)
    for i, w in enumerate([5, 16, 18, 10, 5, 8, 70], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=4):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    def aux_sheet(title, headers, rows):
        s = wb.create_sheet(title)
        s.append(headers)
        for c in s[1]:
            c.font = Font(bold=True)
        for r in rows:
            s.append(r)
        for i, w in enumerate([14, 26, 60][:len(headers)], 1):
            s.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    mons = spec.get("monitors")
    if mons:
        rows = [([m.get("mix", ""), m.get("who", ""), " · ".join(x for x in [m.get("type", ""), m.get("note", "")] if x)]
                 if isinstance(m, dict) else ["", "", str(m)]) for m in mons]
        aux_sheet("Monitors", ["Mix", "Who", "Type / Notes"], rows)
    revs = spec.get("reverbs")
    if revs:
        rows = []
        for r in revs:
            if isinstance(r, dict):
                rows.append([r.get("role", r.get("bus", "")), r.get("preset", ""),
                             r.get("settings", ""), r.get("plugin_eq", ""),
                             " · ".join(x for x in [r.get("why", ""), r.get("note", "")] if x)])
            else:
                rows.append(["", str(r), "", "", ""])
        if spec.get("reverb_pairing"):
            rows.append(["TOGETHER", "", "", "", spec["reverb_pairing"]])
        s = wb.create_sheet("Reverbs")
        s.append(["Role / Bus", "Preset (verbatim from KB)", "Settings", "In-plugin EQ", "Why / How to use"])
        for c in s[1]:
            c.font = Font(bold=True)
        for r in rows:
            s.append(r)
        for i, w in enumerate([12, 30, 34, 30, 60], 1):
            s.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in s.iter_rows(min_row=2):
            for c in row[2:]:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    p = os.path.join(folder, f"{show} - Input List.xlsx"); wb.save(p); return p

# ---------------------------------------------------------------- 3. Show Packet PDF
def build_packet_pdf(spec, folder, packet_builder_path):
    s = importlib.util.spec_from_file_location("pb", packet_builder_path)
    pb = importlib.util.module_from_spec(s); s.loader.exec_module(pb)
    accent = {"DRUMS": pb.DRUMS_BAR, "RHYTHM": pb.GUITAR_BAR, "BASS": pb.BASS_BAR,
              "GUITAR": pb.GUITAR_BAR, "KEYS": pb.KEYS_BAR, "PIANO": pb.PIANO_BAR,
              "STRINGS": pb.STRINGS_BAR, "HORNS": pb.HORNS_BAR, "VOCALS": pb.VOCALS_BAR,
              "AMBIENT": pb.AMBIENT_BAR}
    il, eqc = [], []
    for ch in spec["channels"]:
        sec = ch.get("section", "SPARE")
        il.append(dict(ch=ch["ch"], name=ch["name"], mic=ch["mic"], section=sec,
                       patch=patch_label(ch), phantom_48v=bool(ch.get("phantom")),
                       stand=ch.get("stand", "—"), notes=flag_notes(ch) or ""))
        rows = [dict(band="LC", freq=fmt_freq(ch.get("hpf", 20)), type="LC", gain="—",
                     q="—", notes="High-pass")]
        bb = bands_by_num(ch)
        for n in (4, 3, 2, 1):
            b = bb[n]
            if not b:
                rows.append(dict(band=str(n), freq="—", type="OFF", gain="—", q="—", notes="Flat")); continue
            note = ""
            d = b.get("deq")
            if d:
                note = f"DYNAMIC: thr={d['thr']:g}dB atk={d['atk_ms']:g}ms rel={d['rel_ms']:g}ms — bites only on peaks"
            rows.append(dict(band=str(n), freq=fmt_freq(b["freq"]),
                             type=("Shelf" if str(b["type"]).upper() == "SHELF" else "Bell"),
                             gain=f"{b['gain']:+g} dB", q=f"{b['q']:g}", notes=note))
        lpf = ch.get("lpf")
        rows.append(dict(band="HC", freq=("—" if not lpf else fmt_freq(lpf)), type="HC",
                         gain="—", q="—", notes=("No LPF" if not lpf else "Low-pass")))
        # Numeric feed for the packet's per-input EQ response card — the spec's own
        # values, never re-parsed from the display strings above.
        hpf_n = ch.get("hpf", 20)
        curve = dict(hpf=(float(hpf_n) if hpf_n and float(hpf_n) > 20 else None),
                     lpf=(float(lpf) if lpf else None),
                     bands=[dict(b=n, gain=float(bb[n]["gain"]), freq=float(bb[n]["freq"]),
                                 q=float(bb[n]["q"]),
                                 shelf=str(bb[n]["type"]).upper() == "SHELF",
                                 deq=bool(bb[n].get("deq")))
                            for n in (4, 3, 2, 1) if bb[n]])
        eqc.append(dict(ch=ch["ch"], name=ch["name"], mic=ch["mic"], section=sec,
                        accent=accent.get(sec, pb.TBD_BAR), mic_notes=ch.get("mic_notes", ""),
                        bands=rows, curve=curve, summary=ch.get("eq_summary", "")))
    console = "digico" if "225" in spec.get("console_label", "") or "digico" in spec.get("console_label","").lower() else "wing"
    show_data = dict(show_name=spec["show_name"], venue=venue_label(spec) or spec.get("venue", ""),
                     date=spec.get("show_date", ""), console_label=spec.get("console_label", "DiGiCo Quantum 225"),
                     foh_engineer=spec.get("foh_engineer", "Brian Lloyd"),
                     mon_engineer=spec.get("mon_engineer", "TBD"), show_time=spec.get("show_time", "TBD"),
                     rev=spec.get("rev", "Rev 1.0"), channel_count=len(spec["channels"]),
                     style_note=spec.get("style_note", spec.get("room_context", "")))
    p = os.path.join(folder, f"{spec['show_name']} - Show Packet.pdf")
    marks = pb.build_show_packet(p, show_data, il, eqc, console=console)
    return p, (marks or [])

# ---------------------------------------------------------------- 4. EQ Rationale PDF
def build_rationale_pdf(spec, folder):
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                                     KeepTogether, HRFlowable, PageBreak, Flowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    marks = []

    class PageMark(Flowable):
        """Zero-size page recorder — feeds the MASTER's quick-links page."""
        def __init__(self, kind, label, **meta):
            Flowable.__init__(self)
            self.kind, self.label, self.meta = kind, label, meta
            self.width = self.height = 0

        def wrap(self, aw, ah):
            return (0, 0)

        def draw(self):
            marks.append(dict(kind=self.kind, label=self.label,
                              page=self.canv.getPageNumber(), **self.meta))
    HEADER_BG = colors.HexColor("#1A1A2E"); SUBHEAD = colors.HexColor("#0F3460")
    ACCENT = colors.HexColor("#E94560"); CHANGE = colors.HexColor("#FFF3CD")
    DECIDE = colors.HexColor("#E8F0E8")
    SECCOL = {"DRUMS": colors.HexColor("#D4E8D4"), "BASS": colors.HexColor("#D4D4E8"),
              "RHYTHM": colors.HexColor("#E8E4D4"), "GUITAR": colors.HexColor("#E8E4D4"),
              "KEYS": colors.HexColor("#E8D4E8"), "PIANO": colors.HexColor("#E8D4E8"),
              "STRINGS": colors.HexColor("#E8EEF7"), "HORNS": colors.HexColor("#FFE4B5"),
              "VOCALS": colors.HexColor("#E8D4D4"), "AMBIENT": colors.HexColor("#C7D2FE")}
    st = getSampleStyleSheet()
    H1 = ParagraphStyle('H1', parent=st['Title'], textColor=colors.white, fontSize=20, leading=24, alignment=TA_LEFT)
    SUB = ParagraphStyle('SUB', parent=st['Normal'], textColor=colors.white, fontSize=10, leading=13)
    SEC = ParagraphStyle('SEC', parent=st['Heading2'], textColor=colors.white, fontSize=12, leading=14)
    CHH = ParagraphStyle('CHH', parent=st['Heading3'], textColor=HEADER_BG, fontSize=10.5, leading=12, spaceAfter=1)
    EQ = ParagraphStyle('EQ', parent=st['Normal'], fontName='Courier-Bold', fontSize=8.2, leading=10, textColor=SUBHEAD)
    WHY = ParagraphStyle('WHY', parent=st['Normal'], fontSize=8.6, leading=11, textColor=colors.HexColor("#222222"))
    BODY = ParagraphStyle('BODY', parent=st['Normal'], fontSize=9, leading=12.5)
    p = os.path.join(folder, f"{spec['show_name']} - FOH EQ Reasoning.pdf")
    doc = SimpleDocTemplate(p, pagesize=landscape(letter), leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.45*inch, bottomMargin=0.45*inch)
    W = doc.width; els = []
    banner = Table([[Paragraph(f"{esc(spec['show_name'])} — FOH EQ Rationale", H1)],
                    [Paragraph(f"{esc(venue_label(spec) or spec.get('venue',''))} &middot; {esc(spec.get('console_label',''))} &middot; {esc(spec.get('show_date',''))} &middot; {esc(spec.get('rev',''))}", SUB)]], colWidths=[W])
    banner.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),HEADER_BG),('LEFTPADDING',(0,0),(-1,-1),12),
                                ('RIGHTPADDING',(0,0),(-1,-1),12),('TOPPADDING',(0,0),(0,0),10),
                                ('BOTTOMPADDING',(0,-1),(-1,-1),9),('TOPPADDING',(0,1),(-1,1),0)]))
    els += [PageMark("doc", "EQ Rationale"), banner, Spacer(1,8)]

    # --- context: one titled box per topic, prose broken at its enumerated turns
    # (was a single run-on paragraph carrying all of it — Brian, 2026-07-27).
    HDR = ParagraphStyle('HDR', parent=st['Normal'], fontName='Helvetica-Bold', fontSize=9,
                         leading=11, textColor=colors.white)
    RSH = ParagraphStyle('RSH', parent=st['Normal'], fontName='Helvetica-Bold', fontSize=7.6,
                         leading=9.5, textColor=colors.white)
    RSB = ParagraphStyle('RSB', parent=st['Normal'], fontSize=7.6, leading=9.8,
                         textColor=colors.HexColor("#222222"))
    RSC = ParagraphStyle('RSC', parent=RSB, alignment=TA_CENTER, fontName='Helvetica-Bold',
                         fontSize=7, leading=8.5)

    def titled_box(title, body_html, bg, border=SUBHEAD):
        t = Table([[Paragraph(esc(title).upper(), HDR)], [Paragraph(body_html, WHY)]],
                  colWidths=[W])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),border),
                               ('BACKGROUND',(0,1),(-1,1),bg),
                               ('BOX',(0,0),(-1,-1),0.5,border),
                               ('LEFTPADDING',(0,0),(-1,-1),9),('RIGHTPADDING',(0,0),(-1,-1),9),
                               ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
        return t

    if spec.get("artist_profile"):
        els += [titled_box("The artist — and what it means for this mix",
                           prose_html(spec["artist_profile"]), colors.HexColor("#F4F0E8")),
                Spacer(1,6)]
    if spec.get("room_context"):
        els += [titled_box("The room — and the conditions on the night",
                           prose_html(spec["room_context"]), colors.HexColor("#EAF2FA")),
                Spacer(1,6)]
    if spec.get("monitors"):
        ml = " &nbsp;&middot;&nbsp; ".join(
            (f"{esc(m.get('mix',''))} {esc(m.get('who',''))} ({esc(m.get('type',''))})".strip()
             if isinstance(m, dict) else esc(str(m))) for m in spec["monitors"])
        els += [titled_box("Monitors", ml, colors.HexColor("#F6F6F6")), Spacer(1,6)]
    if spec.get("reverb_note"):
        els += [titled_box("Reverb note", prose_html(spec["reverb_note"]),
                           colors.HexColor("#EAF2FA")), Spacer(1,6)]

    def note_box(title, items, bg, border):
        rows = [[Paragraph(f"<b>{title}</b>", CHH)]]
        for c in items:
            rows.append([Paragraph(f"&bull; {esc(c)}", WHY)])
        box = Table(rows, colWidths=[W])
        box.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),bg),('BOX',(0,0),(-1,-1),0.5,border),
                                 ('LEFTPADDING',(0,0),(-1,-1),9),('RIGHTPADDING',(0,0),(-1,-1),9),
                                 ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        return box
    if spec.get("changes"):
        els += [KeepTogether([PageMark("doc", "What changed — and why"),
                              note_box("What changed from the KB default / prior rev — and why",
                                       spec["changes"], CHANGE, ACCENT)]), Spacer(1,8)]
    if spec.get("decisions"):
        els += [KeepTogether([PageMark("doc", "Question round / decisions"),
                              note_box("Question round — what was asked, what Brian decided",
                                       spec["decisions"], DECIDE, SUBHEAD)]), Spacer(1,8)]
    # Reverb block — Seventh Heaven Pro suggestions (required every show):
    # preset (verbatim), settings, in-plugin EQ, why, and the pairing note.
    revs = spec.get("reverbs")
    if revs:
        items = []
        for r in revs:
            if isinstance(r, dict):
                seg = (f"<b>{esc(str(r.get('role', r.get('bus',''))).upper())}</b> — "
                       f"<b>{esc(r.get('preset',''))}</b>")
                for k, lbl in (("settings", "Settings"), ("plugin_eq", "In-plugin EQ"),
                               ("why", "Why"), ("note", "Note")):
                    if r.get(k):
                        seg += f" &nbsp;|&nbsp; {lbl}: {esc(r[k])}"
                items.append(seg)
            else:
                items.append(esc(str(r)))
        if spec.get("reverb_pairing"):
            items.append(f"<b>USING THEM TOGETHER</b> — {esc(spec['reverb_pairing'])}")
        rows = [[Paragraph("<b>Reverb — Seventh Heaven Pro (100% wet returns)</b>", CHH)]]
        for it in items:
            rows.append([Paragraph(f"&bull; {it}", WHY)])
        box = Table(rows, colWidths=[W])
        box.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor("#EAF2FA")),
                                 ('BOX',(0,0),(-1,-1),0.5,SUBHEAD),
                                 ('LEFTPADDING',(0,0),(-1,-1),9),('RIGHTPADDING',(0,0),(-1,-1),9),
                                 ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        els += [KeepTogether([PageMark("doc", "Reverb — Seventh Heaven"), box]), Spacer(1,8)]

    # ---------------------------------------------------------------- RESEARCH
    # Its own section, formatted to be READ (Brian, 2026-07-27): the framing
    # facts as three boxes, then one table row per researched unit — capsule
    # fact + external source, the AGREE/DISAGREE/THIN verdict, and the five
    # TRACE layers on their own lines. A legacy free-text research_summary is
    # chunked into per-unit entries instead of one wall of prose.
    def research_section():
        R = spec.get("research")
        out = []
        if not R and not spec.get("research_summary"):
            return out
        out += [PageBreak(), PageMark("doc", "Research & Method"),
                section_bar_x("RESEARCH — WHAT WAS LOOKED UP, AND WHAT IT CHANGED"), Spacer(1,5)]
        if isinstance(R, dict):
            heads = [("Genre — verified first, with named evidence", R.get("genre_verified")),
                     ("The gig", R.get("gig")),
                     ("Conditions (fetched, never assumed)", R.get("conditions"))]
            heads = [(t, v) for t, v in heads if v]
            if heads:
                cells = [[Paragraph(esc(t).upper(), RSH)] for t, _ in heads]
                bodies = [[Paragraph(prose_html(v), RSB)] for _, v in heads]
                cw = [W/len(heads)]*len(heads)
                grid = Table([[c[0] for c in cells], [b[0] for b in bodies]], colWidths=cw)
                grid.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),SUBHEAD),
                    ('BACKGROUND',(0,1),(-1,1),colors.HexColor("#F7F9FC")),
                    ('BOX',(0,0),(-1,-1),0.5,SUBHEAD),
                    ('INNERGRID',(0,0),(-1,-1),0.4,colors.HexColor("#B8C6D8")),
                    ('VALIGN',(0,0),(-1,-1),'TOP'),
                    ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                    ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
                out += [grid, Spacer(1,7)]
            units = R.get("units") or []
            if units:
                hdrs = ["CH", "SOURCE / MIC", "WHAT THE RESEARCH FOUND", "V", "TRACE — base · equip · genre · artist · venue"]
                data = [[Paragraph(h, RSH) for h in hdrs]]
                vstyle = []
                for i, u in enumerate(units, 1):
                    src = f"<b>{esc(u.get('source',''))}</b>"
                    if u.get("mic"):
                        src += f"<br/>{esc(u['mic'])}"
                    found = esc(u.get("finding", ""))
                    if u.get("sources"):
                        found += f"<br/><font color='#555555'><i>{esc(u['sources'])}</i></font>"
                    tr = u.get("trace")
                    if isinstance(tr, dict):
                        tl = [f"<b>{k}</b> &nbsp;{esc(tr[k])}" for k in TRACE_LAYERS if tr.get(k)]
                        tl += [f"<b>{esc(k)}</b> &nbsp;{esc(v)}" for k, v in tr.items()
                               if k not in TRACE_LAYERS]
                        trace = "<br/>".join(tl)
                    else:
                        trace = trace_html(tr or "")
                    verdict = str(u.get("verdict", "")).upper()[:8]
                    data.append([Paragraph(esc(u.get("ch", "")), RSC), Paragraph(src, RSB),
                                 Paragraph(found, RSB), Paragraph(esc(verdict), RSC),
                                 Paragraph(trace, RSB)])
                    vbg = {"AGREE": "#DCEFD8", "DISAGREE": "#FFE9B8",
                           "THIN": "#E6E4EC"}.get(verdict, "#F0F0F0")
                    vstyle.append(('BACKGROUND', (3, i), (3, i), colors.HexColor(vbg)))
                    if i % 2 == 0:
                        vstyle.append(('BACKGROUND', (0, i), (2, i), colors.HexColor("#F7F7F9")))
                        vstyle.append(('BACKGROUND', (4, i), (4, i), colors.HexColor("#F7F7F9")))
                t = Table(data, colWidths=[W*x for x in (0.045, 0.155, 0.325, 0.085, 0.39)],
                          repeatRows=1)
                t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),SUBHEAD),
                                       ('GRID',(0,0),(-1,-1),0.4,colors.HexColor("#B8B8C0")),
                                       ('VALIGN',(0,0),(-1,-1),'TOP'),
                                       ('LEFTPADDING',(0,0),(-1,-1),5),('RIGHTPADDING',(0,0),(-1,-1),5),
                                       ('TOPPADDING',(0,0),(-1,-1),3),
                                       ('BOTTOMPADDING',(0,0),(-1,-1),4)] + vstyle))
                out += [t, Spacer(1,7)]
            rec = R.get("reconciliation")
            if rec:
                items = rec if isinstance(rec, list) else [rec]
                out += [note_box("Reconciliation — every web vs KB fork, and how it was settled",
                                 items, CHANGE, ACCENT), Spacer(1,6)]
            if R.get("kb_writeback"):
                out += [note_box("KB write-back candidates — no article covers these yet",
                                 R["kb_writeback"], DECIDE, SUBHEAD), Spacer(1,6)]
        else:
            chunks = research_chunks(spec.get("research_summary"))
            rows, style = [], []
            for i, (head, mic, verdict, body) in enumerate(chunks):
                cell = f"<b>{esc(head)}</b>" if head else ""
                if mic:
                    cell += f"<br/>{esc(mic)}"
                rows.append([Paragraph(cell, RSB), Paragraph(esc(verdict), RSC),
                             Paragraph(body, RSB)])
                vbg = {"AGREE": "#DCEFD8", "DISAGREE": "#FFE9B8",
                       "THIN": "#E6E4EC"}.get(verdict)
                if vbg:
                    style.append(('BACKGROUND', (1, i), (1, i), colors.HexColor(vbg)))
                if i % 2:
                    style.append(('BACKGROUND', (0, i), (0, i), colors.HexColor("#F7F7F9")))
                    style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor("#F7F7F9")))
            t = Table(rows, colWidths=[W*0.155, W*0.085, W*0.76])
            t.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.4,colors.HexColor("#B8B8C0")),
                                   ('VALIGN',(0,0),(-1,-1),'TOP'),
                                   ('LEFTPADDING',(0,0),(-1,-1),5),('RIGHTPADDING',(0,0),(-1,-1),5),
                                   ('TOPPADDING',(0,0),(-1,-1),3),
                                   ('BOTTOMPADDING',(0,0),(-1,-1),4)] + style))
            out += [t, Spacer(1,7)]
        return out

    def section_bar_x(title):
        t = Table([[Paragraph(title, SEC)]], colWidths=[W])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),SUBHEAD),('LEFTPADDING',(0,0),(-1,-1),10),
                               ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        return t

    _rs = research_section()
    if _rs:
        els += _rs + [PageBreak()]

    def section_bar(title):
        t = Table([[Paragraph(title, SEC)]], colWidths=[W])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),SUBHEAD),('LEFTPADDING',(0,0),(-1,-1),10),
                               ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        return t
    def chan_block(ch):
        rowcol = SECCOL.get(ch.get("section","").upper(), colors.HexColor("#EEEEEE"))
        bb = bands_by_num(ch); parts = []
        for n in (4,3,2,1):
            b = bb[n]
            if not b: parts.append(f"B{n} flat"); continue
            seg = f"B{n} {b['gain']:+g}@{fmt_freq(b['freq'])} Q{b['q']:g} {('SHELF' if str(b['type']).upper()=='SHELF' else 'BELL')}"
            if b.get("deq"): seg += " +DEQ"
            parts.append(seg)
        lpf = ch.get("lpf")
        eqline = f"HPF {ch.get('hpf',20):g} | LPF {'off' if not lpf else f'{lpf:g}'} || " + " &nbsp; ".join(parts)
        head = f"<b>Ch {ch['ch']} &nbsp;|&nbsp; {esc(ch['name'])}</b> &middot; {esc(ch['mic'])}"
        if ch.get("ribbon"): head += f" &nbsp;<font color='#9B2222'><b>{RIBBON_FLAG}</b></font>"
        if ch.get("tour"): head += f" &nbsp;<font color='#8a6d00'><b>{TOUR_FLAG}</b></font>"
        inner = [[Paragraph(head, CHH)], [Paragraph(eqline, EQ)]]
        if ch.get("mic_notes"): inner.append([Paragraph(f"<i>{esc(ch['mic_notes'])}</i>", WHY)])
        if ch.get("eq_summary"): inner.append([Paragraph(esc(ch["eq_summary"]), WHY)])
        t = Table(inner, colWidths=[W])
        t.setStyle(TableStyle([('LEFTPADDING',(0,0),(-1,-1),9),('RIGHTPADDING',(0,0),(-1,-1),9),
                               ('TOPPADDING',(0,0),(0,0),4),('BOTTOMPADDING',(0,-1),(-1,-1),5),
                               ('TOPPADDING',(0,1),(-1,-1),1),('BACKGROUND',(0,0),(-1,-1),rowcol),
                               ('LINEBELOW',(0,-1),(-1,-1),0.4,colors.HexColor("#cccccc"))]))
        return KeepTogether([t, Spacer(1,3)])
    last_sec = None
    for ch in spec["channels"]:
        sec = ch.get("section","").upper()
        if sec != last_sec:
            els += [KeepTogether([PageMark("rat_section", sec), section_bar(sec), Spacer(1,3)])]
            last_sec = sec
        els.append(chan_block(ch))
    els += [Spacer(1,6), HRFlowable(width="100%", thickness=0.5, color=ACCENT), Spacer(1,3),
            Paragraph("<font size=7 color='#777777'>Deep-research pass — values reasoned from mic behavior, instrument, genre, the artist's references, and the room. Informed starting points, not gospel; trust your ears at soundcheck.</font>", BODY)]
    doc.build(els); return p, marks

# ---------------------------------------------------------------- 5. MASTER PDF
SEC_CHIP = {"DRUMS": "#FDE68A", "BASS": "#D9CBB5", "RHYTHM": "#BBF7D0", "GUITAR": "#BBF7D0",
            "KEYS": "#DDD6FE", "PIANO": "#FBCFE8", "STRINGS": "#BFDBFE", "HORNS": "#FCD9B4",
            "VOCALS": "#DDD6FE", "AMBIENT": "#C7D2FE", "SPARE": "#E5E7EB"}


def build_nav_pdf(spec, path, groups, console="digico"):
    """Draw the MASTER's QUICK LINKS page(s).

    groups: [(group_title, [entry, ...]), ...] where entry is
      {"label", "sub", "page" (0-based MASTER page), "style": wide|bar|chip, "color"}
    Returns (page_count, links) with links = [(nav_page_idx, (x0,y0,x1,y1), target_page)].
    Coordinates are PDF user space, so pypdf can hang Link annotations straight off them.
    """
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfbase.pdfmetrics import stringWidth

    PW, PH = letter
    M = 0.5 * inch
    X0, X1 = M, PW - M
    W = X1 - X0
    DARK = colors.HexColor("#1A3A5C" if console == "digico" else "#1A1A1A")
    ACC = colors.HexColor("#2E6DA4" if console == "digico" else "#9B2222")
    GREY = colors.HexColor("#6B7280")
    c = rl_canvas.Canvas(path, pagesize=letter)
    st = {"page": 0, "y": PH - M}
    links = []

    def banner(first):
        y = st["y"]
        h = 52 if first else 26
        c.setFillColor(DARK); c.rect(X0, y - h, W, h, stroke=0, fill=1)
        c.setFillColor(colors.white)
        if first:
            c.setFont("Helvetica-Bold", 20)
            c.drawString(X0 + 12, y - 26, f"{spec['show_name']} — QUICK LINKS")
            c.setFont("Helvetica", 9.5)
            sub = " · ".join(x for x in [venue_label(spec) or spec.get("venue", ""),
                                         spec.get("console_label", ""), spec.get("show_date", ""),
                                         spec.get("rev", "")] if x)
            c.drawString(X0 + 12, y - 42, sub)
        else:
            c.setFont("Helvetica-Bold", 11)
            c.drawString(X0 + 12, y - 18, f"{spec['show_name']} — QUICK LINKS (cont.)")
        st["y"] = y - h - (12 if first else 8)
        if first:
            c.setFillColor(GREY); c.setFont("Helvetica-Oblique", 8.5)
            c.drawString(X0 + 2, st["y"] - 2,
                         "Click any row to jump. Page numbers are MASTER pages — "
                         "the reader's bookmark sidebar carries the same map.")
            st["y"] -= 14

    def newpage():
        c.showPage(); st["page"] += 1; st["y"] = PH - M
        banner(False)

    def need(h):
        if st["y"] - h < M + 10:
            newpage()

    def link(rect, page):
        links.append((st["page"], rect, page))

    def group_title(txt):
        need(26)
        y = st["y"] - 16
        c.setFillColor(ACC); c.rect(X0, y, W, 14, stroke=0, fill=1)
        c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 9)
        c.drawString(X0 + 8, y + 4, txt.upper())
        st["y"] = y - 5

    def leader(x_from, x_to, y):
        c.setFillColor(colors.HexColor("#B9C2CC")); c.setFont("Helvetica", 8)
        dots = ""
        while stringWidth(dots + ".", "Helvetica", 8) < (x_to - x_from):
            dots += "."
        if dots:
            c.drawString(x_from, y, dots)

    def wide_row(e):
        h = 20
        need(h)
        y = st["y"] - h
        c.setFillColor(colors.HexColor(e.get("color", "#F1F5F9")))
        c.rect(X0, y + 2, W, h - 3, stroke=0, fill=1)
        c.setFillColor(colors.black); c.setFont("Helvetica-Bold", 10)
        c.drawString(X0 + 9, y + 8, e["label"])
        lx = X0 + 9 + stringWidth(e["label"], "Helvetica-Bold", 10) + 8
        if e.get("sub"):
            c.setFillColor(GREY); c.setFont("Helvetica", 8)
            c.drawString(lx, y + 8, e["sub"])
            lx += stringWidth(e["sub"], "Helvetica", 8) + 8
        pg = f"p {e['page'] + 1}"
        px = X1 - 9 - stringWidth(pg, "Helvetica-Bold", 9)
        leader(lx, px - 6, y + 8)
        c.setFillColor(DARK); c.setFont("Helvetica-Bold", 9)
        c.drawString(px, y + 8, pg)
        link((X0, y + 2, X1, y + h - 1), e["page"])
        st["y"] = y

    def bar_row(e):
        h = 15
        need(h + 4)
        y = st["y"] - h
        c.setFillColor(colors.HexColor(e.get("color", "#E5E7EB")))
        c.rect(X0, y, W, h - 2, stroke=0, fill=1)
        c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 8.5)
        c.drawString(X0 + 8, y + 4, e["label"].upper())
        c.setFont("Helvetica-Bold", 8)
        c.drawRightString(X1 - 8, y + 4, f"p {e['page'] + 1}")
        link((X0, y, X1, y + h - 2), e["page"])
        st["y"] = y - 2

    def chips(entries):
        cols, gap = 3, 8
        cw = (W - gap * (cols - 1)) / cols
        h = 15
        for i in range(0, len(entries), cols):
            row = entries[i:i + cols]
            need(h + 2)
            y = st["y"] - h
            for j, e in enumerate(row):
                x = X0 + j * (cw + gap)
                c.setFillColor(colors.HexColor(e.get("color", "#F3F4F6")))
                c.rect(x, y, cw, h - 2, stroke=0, fill=1)
                c.setFillColor(colors.HexColor("#111827")); c.setFont("Helvetica-Bold", 8)
                lab = e["label"]
                maxw = cw - 34
                while stringWidth(lab, "Helvetica-Bold", 8) > maxw and len(lab) > 4:
                    lab = lab[:-2]
                c.drawString(x + 6, y + 4, lab)
                c.setFillColor(DARK); c.setFont("Helvetica", 7.5)
                c.drawRightString(x + cw - 5, y + 4, f"p{e['page'] + 1}")
                link((x, y, x + cw, y + h - 2), e["page"])
            st["y"] = y - 2
        st["y"] -= 3

    banner(True)
    for title, entries in groups:
        if not entries:
            continue
        group_title(title)
        buf = []
        for e in entries:
            style = e.get("style", "wide")
            if style == "chip":
                buf.append(e); continue
            if buf:
                chips(buf); buf = []
            (bar_row if style == "bar" else wide_row)(e)
        if buf:
            chips(buf)
        st["y"] -= 6
    c.save()
    return st["page"] + 1, links


def build_master_pdf(spec, folder, parts):
    """One giant PDF with everything (Brian, 2026-07-08), now opening on a
    clickable QUICK LINKS page (2026-07-27): jump straight to the input list,
    any EQ channel, the research section, or the stage plot instead of scrolling.

    parts: [(pdf_path, page_marks), ...] in MASTER order. Band-provided
    "<Show> - Stage Plot.pdf" / "- Rider.pdf" in the folder are appended.
    The individual files still ship — this is additive."""
    from pypdf import PdfWriter, PdfReader
    from pypdf.annotations import Link
    from pypdf.generic import ArrayObject, FloatObject, NameObject, NumberObject

    show = spec["show_name"]
    parts = [(p, list(m or [])) for p, m in parts]
    for tail, label in ((" - Stage Plot.pdf", "Stage Plot"), (" - Rider.pdf", "Rider")):
        extra = os.path.join(folder, f"{show}{tail}")
        if os.path.exists(extra):
            parts.append((extra, [dict(kind="doc", label=label, page=1)]))
    counts = [len(PdfReader(p).pages) for p, _ in parts]
    console = "digico" if ("225" in spec.get("console_label", "")
                           or "digico" in spec.get("console_label", "").lower()) else "wing"
    nav_path = os.path.join(folder, f".{show} - navtmp.pdf")

    def collect(nav_pages):
        """Group the page marks into nav rows, offset by the nav page count."""
        offs, run = [], nav_pages
        for n in counts:
            offs.append(run); run += n
        docs, sections, rat = [], {}, []
        order = []
        for (path, marks), off in zip(parts, offs):
            for m in marks:
                pg = off + int(m.get("page", 1)) - 1
                kind = m.get("kind")
                if kind == "doc":
                    docs.append(dict(label=m["label"], page=pg, style="wide",
                                     color="#E8EEF7"))
                elif kind == "section":
                    sec = m["label"].upper()
                    sections.setdefault(sec, {"page": pg, "chans": []})
                    order.append(sec)
                elif kind == "channel":
                    sec = str(m.get("section", "")).upper()
                    sections.setdefault(sec, {"page": pg, "chans": []})["chans"].append(
                        dict(label=f"{m.get('ch','')}  {m['label']}", page=pg, style="chip",
                             color=SEC_CHIP.get(sec, "#F3F4F6")))
                elif kind == "rat_section":
                    rat.append(dict(label=m["label"], page=pg, style="chip",
                                    color=SEC_CHIP.get(m["label"].upper(), "#F3F4F6")))
        eq = []
        seen = []
        for sec in order:
            if sec in seen:
                continue
            seen.append(sec)
            d = sections[sec]
            eq.append(dict(label=sec, page=d["page"], style="bar",
                           color=SEC_CHIP.get(sec, "#E5E7EB")))
            eq += d["chans"]
        groups = [("Documents", docs),
                  ("EQ pages — by section and channel (Show Packet)", eq),
                  ("EQ rationale — by section", rat)]
        return groups

    nav_pages, links = 1, []
    for _ in range(4):                      # settle: nav length shifts every offset
        groups = collect(nav_pages)
        n, links = build_nav_pdf(spec, nav_path, groups, console=console)
        if n == nav_pages:
            break
        nav_pages = n
    else:
        groups = collect(nav_pages)

    w = PdfWriter()
    w.append(nav_path)
    for p, _ in parts:
        w.append(p)
    # Clickable rows on the quick-links page(s)
    for nav_pg, rect, target in links:
        try:
            w.add_annotation(page_number=nav_pg,
                             annotation=Link(rect=rect, target_page_index=target))
        except Exception as exc:                       # never lose the MASTER over a link
            print(f"  note: quick-link skipped ({exc})")
    # Kill the visible link borders some viewers draw
    for pg in range(nav_pages):
        for annot in (w.pages[pg].get("/Annots") or []):
            obj = annot.get_object()
            if obj.get("/Subtype") == "/Link":
                obj[NameObject("/Border")] = ArrayObject(
                    [NumberObject(0), NumberObject(0), NumberObject(0)])
                obj[NameObject("/C")] = ArrayObject(
                    [FloatObject(1), FloatObject(1), FloatObject(1)])
    # Same map as PDF bookmarks, for the reader's sidebar
    try:
        for title, entries in groups:
            if not entries:
                continue
            parent = w.add_outline_item(title, entries[0]["page"])
            sub = None
            for e in entries:
                if e.get("style") == "bar":
                    sub = w.add_outline_item(e["label"], e["page"], parent=parent)
                else:
                    w.add_outline_item(e["label"], e["page"], parent=sub or parent)
    except Exception as exc:
        print(f"  note: bookmarks skipped ({exc})")
    out = os.path.join(folder, f"{show} - MASTER.pdf")
    with open(out, "wb") as f:
        w.write(f)
    if os.path.exists(nav_path):
        os.remove(nav_path)
    return out

# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--spec", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--packet-builder",
                    default=os.path.expanduser("~/Documents/Claude/audio/show-packet-builder-template.py"))
    a = ap.parse_args()
    spec = load_spec(a.spec)
    errors, warnings = validate_spec(spec)
    for w in warnings:
        print("  spec warn:", w)
    if errors:
        for e in errors:
            print("  spec ERROR:", e)
        raise SystemExit(f"spec validation failed — {len(errors)} error(s), nothing written")
    print(f"spec validation PASS — {len(spec['channels'])} channels, {len(warnings)} warnings")
    os.makedirs(a.out, exist_ok=True)
    md = write_md(spec, a.out)
    lint_md(md)
    packet_pdf, packet_marks = build_packet_pdf(spec, a.out, a.packet_builder)
    rationale_pdf, rat_marks = build_rationale_pdf(spec, a.out)
    made = [md, write_xlsx(spec, a.out), packet_pdf, rationale_pdf,
            build_master_pdf(spec, a.out, [(packet_pdf, packet_marks),
                                           (rationale_pdf, rat_marks)])]
    for m in made:
        print("WROTE", m)
    try:
        sys.path.insert(0, os.path.expanduser("~/Documents/Claude/audio/_shared"))
        import show_status
        show_status.stamp(a.out, "packet_built",
                          note=f"{len(spec['channels'])} channels, full packet")
    except ImportError:
        pass  # status stamp is best-effort — never blocks a build
    print("\nNEXT: run the venue .ses patcher on the .md, then verify on the console.")

if __name__ == "__main__":
    main()
